#!/usr/bin/env python3
"""
TFB symbol de-duplication and identity repair.
=================================================================
Removes duplicate security rows created by inconsistent ticker
suffixes (AAPL vs AAPL.US) WITHOUT destroying distinct companies
that merely share a ticker root (ALV.US=Autoliv vs ALV.DE=Allianz).

WHY THE NAIVE RULE FAILS
------------------------
The obvious rule -- "same name AND same price => delete one" --
only catches ~16% of real duplicates, because the twin rows are
refreshed on different days, so their prices legitimately differ
(NTES 121.11 vs NTES.US 129.55 on the same date). Meanwhile a
base-ticker-only rule would delete real companies.

THE RULE USED HERE  (all three must agree)
------------------------------------------
  1. base ticker  (symbol with exchange suffix stripped)
  2. normalised company name (legal suffixes/punctuation removed)
  3. currency
=> same security. Keep the FRESHEST row, drop the rest.

Anything else is classified, reported, and LEFT ALONE:
  - same base + different name      -> different companies  (NEVER touch)
  - same name + different currency  -> cross-listing         (keep, tag)
  - blank identity ("empty shell")  -> quarantined row       (drop or flag)

Usage:
    python3 tfb_dedup.py input.xlsx --out cleaned.xlsx
    python3 tfb_dedup.py input.xlsx --dry-run
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone

import openpyxl

# --------------------------------------------------------------------------
# configuration
# --------------------------------------------------------------------------

# Sheets that hold one row per security.
DATA_SHEETS = ["Global_Markets", "Market_Leaders", "Mutual_Funds",
               "Commodities_FX", "My_Portfolio"]

# Exchange suffixes we recognise. A trailing token NOT in this list is
# treated as part of the ticker, so we never mangle names like BRK.B.
EXCHANGE_SUFFIXES = {
    "US", "SR", "L", "TO", "V", "NE", "CN", "PA", "DE", "F", "MI", "AS",
    "BR", "LS", "MC", "SW", "VI", "ST", "CO", "HE", "OL", "IR", "WA",
    "PR", "BUD", "AT", "IS", "TA", "SA", "MX", "BA", "SN", "LM", "CR",
    "HK", "SS", "SZ", "T", "KS", "KQ", "TW", "TWO", "BK", "JK", "SI",
    "KL", "NS", "BO", "CM", "AX", "NZ", "VN", "PH", "JO", "KW", "QA",
    "AE", "EG", "MA", "NG", "KE", "ZA", "JO", "TR", "IL", "PSE",
}

# Legal-form tokens stripped when normalising a company name.
NAME_NOISE = {
    "INC", "INCORPORATED", "CORP", "CORPORATION", "COMPANY", "CO", "COS",
    "LTD", "LIMITED", "PLC", "LLC", "LP", "LLP", "AG", "NV", "BV", "SE",
    "SA", "SAB", "SAA", "SPA", "AB", "AS", "ASA", "OYJ", "KK", "KGAA",
    "GMBH", "HOLDING", "HOLDINGS", "GROUP", "GRP", "THE", "AND",
    "SGPS", "CV", "DE", "CIA", "COMPANHIA", "SAS", "PT", "TBK", "BHD",
    "PJSC", "JSC", "OAO", "PAO", "ADR", "GDR", "CLASS", "SHARES", "SHS",
    "COMMON", "STOCK", "ORD", "ORDINARY", "REG", "REGISTERED",
}

# Price divergence above this between otherwise-identical twins is
# reported as a staleness warning (it does not change the decision).
PRICE_DIVERGENCE_WARN = 0.02

# --------------------------------------------------------------------------
# identity helpers
# --------------------------------------------------------------------------


def split_symbol(symbol):
    """'AAPL.US' -> ('AAPL', 'US');  'BRK.B' -> ('BRK.B', '')."""
    if symbol is None:
        return "", ""
    s = str(symbol).strip().upper()
    if "." not in s:
        return s, ""
    root, _, tail = s.rpartition(".")
    if tail in EXCHANGE_SUFFIXES and root:
        return root, tail
    return s, ""


def normalise_name(name):
    """Company name -> comparable key. Empty string if unusable."""
    if name is None:
        return ""
    n = str(name).strip().upper()
    if not n:
        return ""
    n = re.sub(r"[^A-Z0-9 ]", " ", n)
    tokens = [t for t in n.split() if t and t not in NAME_NOISE]
    return "".join(tokens)


def parse_when(value):
    """Best-effort timestamp parse -> aware datetime, or None."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    txt = str(value).strip().replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(txt)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def to_float(value):
    try:
        f = float(str(value).strip())
        return f if f == f else None          # reject NaN
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------
# row model
# --------------------------------------------------------------------------


class Row:
    """One security row, with everything needed to judge it."""

    __slots__ = ("index", "excel_row", "symbol", "base", "suffix", "name",
                 "name_key", "currency", "price", "updated", "warnings",
                 "filled")

    def __init__(self, index, excel_row, values, col):
        self.index = index
        self.excel_row = excel_row
        self.symbol = str(values[col["Symbol"]] or "").strip()
        self.base, self.suffix = split_symbol(self.symbol)
        raw_name = values[col["Name"]] if "Name" in col else None
        self.name = str(raw_name or "").strip()
        self.name_key = normalise_name(raw_name)
        self.currency = str(
            values[col["Currency"]] or "").strip().upper() if "Currency" in col else ""
        self.price = to_float(
            values[col["Current Price"]]) if "Current Price" in col else None
        self.updated = None
        for key in ("Last Updated (UTC)", "Last Updated (Riyadh)"):
            if key in col:
                self.updated = self.updated or parse_when(values[col[key]])
        self.warnings = str(
            values[col["Warnings"]] or "") if "Warnings" in col else ""
        # how much real data this row carries -- used to break freshness ties
        self.filled = sum(
            1 for v in values if v is not None and str(v).strip() != "")

    @property
    def is_shell(self):
        """Identity-quarantined row: symbol present, everything else blank."""
        return self.name_key == "" and self.price is None

    def freshness_key(self):
        """Higher sorts better: freshest, then most complete."""
        stamp = self.updated.timestamp() if self.updated else float("-inf")
        return (stamp, self.filled, self.price is not None)

    def __repr__(self):
        return f"<{self.symbol} {self.name[:24]!r} {self.currency} {self.price}>"


# --------------------------------------------------------------------------
# classification
# --------------------------------------------------------------------------


def classify(rows):
    """
    Bucket rows sharing a base ticker.

    Returns (drop, findings) where drop is a set of Row.index to delete
    and findings is a list of dicts describing every decision made.
    """
    drop, findings = set(), []
    by_base = defaultdict(list)
    for row in rows:
        if row.base:
            by_base[row.base].append(row)

    # --- pass 1: exact-identity duplicates (base + name + currency) -------
    by_identity = defaultdict(list)
    for row in rows:
        if row.name_key and row.base:
            by_identity[(row.base, row.name_key, row.currency)].append(row)

    for (base, _nk, ccy), group in by_identity.items():
        if len(group) < 2:
            continue
        group.sort(key=Row.freshness_key, reverse=True)
        keep, losers = group[0], group[1:]
        prices = [r.price for r in group if r.price is not None]
        spread = ((max(prices) - min(prices)) / max(prices)
                  if len(prices) >= 2 and max(prices) else 0.0)
        for loser in losers:
            drop.add(loser.index)
        findings.append({
            "kind": "DUPLICATE_REMOVED",
            "base": base,
            "currency": ccy,
            "company": keep.name,
            "kept": keep.symbol,
            "kept_price": keep.price,
            "kept_updated": keep.updated.isoformat() if keep.updated else "",
            "dropped": [r.symbol for r in losers],
            "dropped_prices": [r.price for r in losers],
            "price_spread": round(spread, 4),
            "note": ("stale twin: prices disagree by "
                     f"{spread:.1%} -- exact-price matching would have MISSED this"
                     if spread > PRICE_DIVERGENCE_WARN else "prices agree"),
        })

    # --- pass 2: everything else that shares a base ticker ----------------
    for base, group in by_base.items():
        live = [r for r in group if r.index not in drop]
        if len(live) < 2:
            continue
        named = [r for r in live if r.name_key]
        if len({r.name_key for r in named}) > 1:
            findings.append({
                "kind": "DISTINCT_COMPANIES_KEPT",
                "base": base,
                "symbols": [r.symbol for r in live],
                "companies": [r.name for r in live],
                "prices": [r.price for r in live],
                "note": "different issuers sharing a ticker root -- NOT merged",
            })
        elif len({r.currency for r in named if r.currency}) > 1:
            findings.append({
                "kind": "CROSS_LISTING_KEPT",
                "base": base,
                "company": named[0].name if named else "",
                "symbols": [r.symbol for r in live],
                "currencies": [r.currency for r in live],
                "prices": [r.price for r in live],
                "note": "same issuer, different venue/currency -- kept, "
                        "but treat as ONE exposure for concentration limits",
            })

    # --- pass 2b: rows displaying a FALLBACK price -------------------------
    # The provider's live quote failed but a number was rendered anyway, with
    # no visual cue. This is how BK -> "Hanwha Aerospace @ 979,000 USD" and
    # BRK-B -> "Taiwan Semiconductor" reach the sheet looking legitimate.
    for row in rows:
        if row.index in drop:
            continue
        if "quote_current_price_missing" in row.warnings and row.price is not None:
            findings.append({
                "kind": "SUSPECT_QUOTE_KEPT",
                "base": row.base,
                "company": row.name,
                "symbols": [row.symbol],
                "currency": row.currency,
                "prices": [row.price],
                "note": "live quote FAILED; displayed price is a fallback -- "
                        "verify identity and price before trading",
            })

    # --- pass 3: identity-quarantined empty shells ------------------------
    for base, group in by_base.items():
        shells = [r for r in group if r.is_shell and r.index not in drop]
        if not shells:
            continue
        live = [r for r in group
                if not r.is_shell and r.index not in drop and r.price is not None]
        for shell in shells:
            drop.add(shell.index)
        if live:
            findings.append({
                "kind": "SHELL_DROPPED_TWIN_ALIVE",
                "base": base,
                "dropped": [r.symbol for r in shells],
                "surviving": [r.symbol for r in live],
                "note": "blanked duplicate; a populated twin remains",
            })
        else:
            findings.append({
                "kind": "SHELL_DROPPED_NEEDS_REFETCH",
                "base": base,
                "dropped": [r.symbol for r in shells],
                "surviving": [],
                "note": "ORPHAN: row was blanked but has no populated twin -- "
                        "data was destroyed, symbol needs re-fetch",
            })
    return drop, findings


# --------------------------------------------------------------------------
# sheet processing
# --------------------------------------------------------------------------


def find_header(worksheet, probe=12):
    """Locate the header row containing 'Symbol'. Returns (row_no, {name: idx})."""
    for row_no, values in enumerate(
            worksheet.iter_rows(min_row=1, max_row=probe, values_only=True), start=1):
        cells = [str(v).strip() if v is not None else "" for v in values]
        if "Symbol" in cells:
            return row_no, {name: i for i, name in enumerate(cells) if name}
    return None, None


def process_sheet(worksheet):
    header_row, col = find_header(worksheet)
    if not col or "Symbol" not in col:
        return None
    rows, sym_idx = [], col["Symbol"]
    for offset, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True)):
        if not values or sym_idx >= len(values):
            continue
        if values[sym_idx] is None or str(values[sym_idx]).strip() == "":
            continue
        rows.append(Row(len(rows), header_row + 1 + offset, values, col))
    drop, findings = classify(rows)
    return {"rows": rows, "drop": drop, "findings": findings,
            "excel_rows_to_delete": sorted(
                (r.excel_row for r in rows if r.index in drop), reverse=True)}


def compact_sheet(worksheet, result):
    """
    Remove the dropped rows by shifting survivors up, then clearing the tail.

    openpyxl's delete_rows() is O(cells) per call, so 700 individual calls on a
    10k-row x 115-col sheet does not finish in reasonable time. Rewriting the
    data block in place is a single pass.
    """
    keep = [r for r in result["rows"] if r.index not in result["drop"]]
    if not keep:
        return
    first_data_row = min(r.excel_row for r in result["rows"])
    last_data_row = max(r.excel_row for r in result["rows"])
    width = worksheet.max_column

    # snapshot survivor values before we start overwriting
    survivors = [[worksheet.cell(row=r.excel_row, column=c).value
                  for c in range(1, width + 1)] for r in keep]

    for offset, values in enumerate(survivors):
        target = first_data_row + offset
        for col_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=target, column=col_index)
            if cell.value != value:
                cell.value = value

    for row_no in range(first_data_row + len(survivors), last_data_row + 1):
        for col_index in range(1, width + 1):
            worksheet.cell(row=row_no, column=col_index).value = None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook")
    ap.add_argument("--out", help="write cleaned workbook here")
    ap.add_argument("--dry-run", action="store_true",
                    help="report only, change nothing")
    ap.add_argument("--report", default="dedup_report.csv")
    ap.add_argument("--sheets", nargs="*", default=DATA_SHEETS)
    args = ap.parse_args()

    if not args.dry_run and not args.out:
        ap.error("give --out or --dry-run")

    print(f"loading {args.workbook} ...")
    wb = openpyxl.load_workbook(args.workbook)

    all_findings, totals = [], {}
    for name in args.sheets:
        if name not in wb.sheetnames:
            print(f"  {name}: absent, skipped")
            continue
        result = process_sheet(wb[name])
        if result is None:
            print(f"  {name}: no Symbol column, skipped")
            continue
        for finding in result["findings"]:
            finding["sheet"] = name
        all_findings += result["findings"]
        counts = defaultdict(int)
        for finding in result["findings"]:
            counts[finding["kind"]] += 1
        totals[name] = {"rows": len(result["rows"]),
                        "removed": len(result["drop"]), **counts}
        print(f"  {name}: {len(result['rows'])} rows -> "
              f"remove {len(result['drop'])}  "
              f"({dict(counts)})")
        if not args.dry_run and result["drop"]:
            compact_sheet(wb[name], result)

    # ---- write the audit trail ----
    import csv as _csv
    fields = ["sheet", "kind", "base", "company", "currency", "kept",
              "kept_price", "kept_updated", "dropped", "dropped_prices",
              "surviving", "symbols", "companies", "currencies", "prices",
              "price_spread", "note"]
    with open(args.report, "w", newline="", encoding="utf-8") as fh:
        writer = _csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        order = {"DUPLICATE_REMOVED": 0, "SHELL_DROPPED_NEEDS_REFETCH": 1,
                 "SHELL_DROPPED_TWIN_ALIVE": 2, "CROSS_LISTING_KEPT": 3,
                 "DISTINCT_COMPANIES_KEPT": 4}
        for finding in sorted(all_findings,
                             key=lambda f: (order.get(f["kind"], 9), f["sheet"],
                                            f.get("base", ""))):
            writer.writerow({k: (", ".join(str(x) for x in v)
                                 if isinstance(v, list) else v)
                             for k, v in finding.items()})
    print(f"\naudit trail -> {args.report}  ({len(all_findings)} findings)")

    if not args.dry_run:
        wb.save(args.out)
        print(f"cleaned workbook -> {args.out}")

    print("\n" + "=" * 62)
    grand = defaultdict(int)
    for stats in totals.values():
        for key, value in stats.items():
            if key != "rows":
                grand[key] += value
    for key, value in sorted(grand.items()):
        print(f"  {key:<32} {value:>6}")
    print("=" * 62)
    return 0


if __name__ == "__main__":
    sys.exit(main())
