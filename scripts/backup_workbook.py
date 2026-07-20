#!/usr/bin/env python3
# scripts/backup_workbook.py
"""
================================================================================
Workbook Backup — v1.0.0 (W-4, Execution Plan v2.1; Gate & Evidence Register §8)
================================================================================
NEW script (window whitelist W-4, scheduled Jul 22–24; adopted 2026-07-20).

WHY
    The production classifier ("Market Share Deepseek-V3") is a single point
    of failure: every decision surface, ledger, and evidence tab lives in one
    Google Sheet with no independent copy. Register §8 sets the contract this
    script implements: RPO ≤ 24h · RTO ≤ 60min via runbook · retention
    30 daily + 12 monthly · restorer = operator · scope = WORKBOOK DATA ONLY.

WHAT IT DOES (nightly via .github/workflows/workbook_backup.yml)
    1. Exports the workbook as full-size xlsx through the Docs export URL
       (https://docs.google.com/spreadsheets/d/{id}/export?format=xlsx).
       DELIBERATE: the Drive API files.export endpoint caps at 10 MB and this
       workbook is ~18 MB — the Docs URL streams the same bytes Google's own
       File▸Download produces, with no cap, authorized by the service
       account's bearer token.
    2. Uploads it to the operator's backup folder as
       DAILY_TFB_Workbook_YYYY-MM-DD.xlsx using a RESUMABLE upload
       (multipart caps at 5 MB — also too small).
    3. On the first run of each month, server-side files.copy of the fresh
       daily into MONTHLY_TFB_Workbook_YYYY-MM.xlsx (no second upload).
    4. Prunes: newest 30 DAILY_ kept, newest 12 MONTHLY_ kept.
    5. Appends an audit line to _Run_Log — [BACKUP v1.0.0] name/size/sha/
       kept-counts on success, an ERROR row on ANY failure — so a silent
       failure is impossible: the seven-gate audit reads _Run_Log daily, and
       the workflow itself goes red (GitHub failure e-mail).

SECURITY SCOPE (Register §8, explicit)
    Workbook DATA only. Credentials, tokens, and ENV configuration are
    EXCLUDED BY DESIGN — secrets live in GitHub/Render and never enter a
    backup artifact; a credential-bearing backup would be a security
    regression, not a safety net. Configuration is recoverable from the repo
    plus the Gate & Evidence Register.

RESTORE RUNBOOK (RTO ≤ 60 min, restorer = operator)
    1. Open the Drive backup folder → download the wanted DAILY_/MONTHLY_
       xlsx (2 min).
    2. sheets.new → File ▸ Import ▸ Upload → "Insert new sheet(s)" into a NEW
       spreadsheet (never overwrite production blind) (5–10 min).
    3. Compare _Status / _Run_Log tails against the incident window; copy the
       needed tabs back into production, or repoint DEFAULT_SPREADSHEET_ID at
       the restored copy if production is unrecoverable (remaining budget).
    4. Verify with the standard shell/version checks + a manual cockpit
       refresh. Post-restore, dispatch this workflow with restore_test=true
       to re-prove the chain.

ONE-TIME OPERATOR SETUP (flagged — GitHub side, NOT Render)
    a. Drive → New folder (e.g. "TFB_Backups").
    b. Share it, role EDITOR, with the service-account e-mail (this script
       PRINTS that e-mail on a folder-access failure so it is never a hunt).
    c. Repo → Settings → Secrets and variables → Actions → Variables →
       New variable: TFB_BACKUP_FOLDER_ID = the folder id from its URL.

ENV
    DEFAULT_SPREADSHEET_ID / SPREADSHEET_ID   production workbook id
    TFB_BACKUP_FOLDER_ID                      Drive folder id (repo Variable)
    GOOGLE_SHEETS_CREDENTIALS(_B64) | GOOGLE_APPLICATION_CREDENTIALS
    TFB_BACKUP_MIN_BYTES     default 1000000  (sanity floor on the export)
    TFB_BACKUP_KEEP_DAILY    default 30
    TFB_BACKUP_KEEP_MONTHLY  default 12

USAGE
    python scripts/backup_workbook.py                 # nightly backup (CI)
    python scripts/backup_workbook.py --dry-run       # export+verify, no upload
    python scripts/backup_workbook.py --restore-test  # acceptance: download
                                                      #   newest daily, open it,
                                                      #   assert core tabs, log
    python scripts/backup_workbook.py --selftest      # offline logic harness
================================================================================
"""
from __future__ import annotations

import argparse
import base64
import datetime as _dt
import hashlib
import io
import json
import os
import re
import sys
import time
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

__version__ = "1.0.0"
_RIYADH = ZoneInfo("Asia/Riyadh")

_XLSX_MIME = ("application/vnd.openxmlformats-officedocument"
              ".spreadsheetml.sheet")
_DAILY_RE = re.compile(r"^DAILY_TFB_Workbook_(\d{4}-\d{2}-\d{2})\.xlsx$")
_MONTHLY_RE = re.compile(r"^MONTHLY_TFB_Workbook_(\d{4}-\d{2})\.xlsx$")
_CORE_TABS = {"_Status", "_Run_Log", "Top_10_Investments",
              "_Portfolio_CostBasis", "My_Portfolio", "Shadow_History"}


def _out(msg: str) -> None:
    sys.stderr.write(f"[workbook_backup v{__version__}] {msg}\n")


def _env(name: str, default: str = "") -> str:
    return (os.getenv(name) or default).strip()


def _env_int(name: str, default: int) -> int:
    try:
        v = int(float(_env(name, str(default))))
        return v if v > 0 else default
    except (TypeError, ValueError):
        return default


def _now_riyadh() -> _dt.datetime:
    return _dt.datetime.now(_RIYADH)


# --------------------------------------------------------------------------- #
# Pure helpers (offline selftest)                                             #
# --------------------------------------------------------------------------- #
def daily_name(d: _dt.date) -> str:
    return f"DAILY_TFB_Workbook_{d.isoformat()}.xlsx"


def monthly_name(d: _dt.date) -> str:
    return f"MONTHLY_TFB_Workbook_{d.strftime('%Y-%m')}.xlsx"


def sha256_hex(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def plan_pruning(names: List[str], keep_daily: int, keep_monthly: int
                 ) -> Tuple[List[str], List[str]]:
    """(keep, delete) over folder file NAMES. Newest-by-embedded-date wins;
    anything not matching either pattern is LEFT ALONE (never delete a file
    this script did not name)."""
    dailies = sorted((n for n in names if _DAILY_RE.match(n)), reverse=True)
    monthlies = sorted((n for n in names if _MONTHLY_RE.match(n)), reverse=True)
    keep = dailies[:keep_daily] + monthlies[:keep_monthly]
    delete = dailies[keep_daily:] + monthlies[keep_monthly:]
    return keep, delete


def monthly_needed(today: _dt.date, existing_names: List[str]) -> bool:
    """A MONTHLY_ copy is due when none exists yet for today's month."""
    want = monthly_name(today)
    return want not in set(existing_names)


# --------------------------------------------------------------------------- #
# Google auth + HTTP (CI-only imports)                                        #
# --------------------------------------------------------------------------- #
def _credentials():
    from google.oauth2 import service_account  # local import: CI only
    scopes = ["https://www.googleapis.com/auth/drive",
              "https://www.googleapis.com/auth/spreadsheets"]
    path = _env("GOOGLE_APPLICATION_CREDENTIALS")
    if path and os.path.exists(path):
        return service_account.Credentials.from_service_account_file(
            path, scopes=scopes)
    raw = (_env("GOOGLE_SHEETS_CREDENTIALS")
           or _env("GOOGLE_SHEETS_CREDENTIALS_B64")
           or _env("GOOGLE_CREDENTIALS"))
    if not raw:
        raise RuntimeError("no Google credentials in env")
    s = raw
    if not s.startswith("{"):
        try:
            dec = base64.b64decode(s).decode("utf-8", errors="replace").strip()
            if dec.startswith("{"):
                s = dec
        except Exception:
            pass
    info = json.loads(s)
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=scopes)
    creds._tfb_client_email = info.get("client_email", "?")  # for error help
    return creds


def _session(creds):
    from google.auth.transport.requests import AuthorizedSession
    return AuthorizedSession(creds)


def export_xlsx(sess, sheet_id: str) -> bytes:
    """Full-size xlsx via the Docs export URL (files.export caps at 10 MB)."""
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    r = sess.get(url, timeout=300)
    r.raise_for_status()
    return r.content


def drive_list(sess, folder_id: str) -> List[Dict[str, Any]]:
    files: List[Dict[str, Any]] = []
    token = None
    while True:
        params = {"q": f"'{folder_id}' in parents and trashed=false",
                  "fields": "nextPageToken,files(id,name,size,createdTime)",
                  "pageSize": 200,
                  "supportsAllDrives": "true",
                  "includeItemsFromAllDrives": "true"}
        if token:
            params["pageToken"] = token
        r = sess.get("https://www.googleapis.com/drive/v3/files",
                     params=params, timeout=60)
        r.raise_for_status()
        j = r.json()
        files += j.get("files", [])
        token = j.get("nextPageToken")
        if not token:
            return files


def drive_upload_resumable(sess, folder_id: str, name: str,
                           data: bytes) -> Dict[str, Any]:
    meta = {"name": name, "parents": [folder_id]}
    r = sess.post(
        "https://www.googleapis.com/upload/drive/v3/files"
        "?uploadType=resumable&supportsAllDrives=true",
        json=meta,
        headers={"X-Upload-Content-Type": _XLSX_MIME,
                 "X-Upload-Content-Length": str(len(data))},
        timeout=60)
    r.raise_for_status()
    loc = r.headers.get("Location")
    if not loc:
        raise RuntimeError("resumable upload: no session Location header")
    r2 = sess.put(loc, data=data,
                  headers={"Content-Type": _XLSX_MIME}, timeout=600)
    r2.raise_for_status()
    return r2.json()


def drive_copy(sess, file_id: str, folder_id: str, new_name: str
               ) -> Dict[str, Any]:
    r = sess.post(
        f"https://www.googleapis.com/drive/v3/files/{file_id}/copy"
        "?supportsAllDrives=true",
        json={"name": new_name, "parents": [folder_id]}, timeout=120)
    r.raise_for_status()
    return r.json()


def drive_delete(sess, file_id: str) -> None:
    r = sess.delete(
        f"https://www.googleapis.com/drive/v3/files/{file_id}"
        "?supportsAllDrives=true", timeout=60)
    if r.status_code not in (200, 204):
        raise RuntimeError(f"delete {file_id}: HTTP {r.status_code}")


def drive_download(sess, file_id: str) -> bytes:
    r = sess.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}"
        "?alt=media&supportsAllDrives=true", timeout=600)
    r.raise_for_status()
    return r.content


# --------------------------------------------------------------------------- #
# _Run_Log audit line (guarded — logging must never mask the real outcome)    #
# --------------------------------------------------------------------------- #
def append_run_log(status: str, message: str, details: Dict[str, Any]) -> None:
    try:
        import gspread
        creds = _credentials()
        gc = gspread.authorize(creds)
        sid = _env("DEFAULT_SPREADSHEET_ID") or _env("SPREADSHEET_ID")
        ws = gc.open_by_key(sid).worksheet("_Run_Log")
        ws.append_row(
            [_dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
             "INFO" if status == "OK" else "ERROR", "workbook_backup",
             "_Backup", status, message, "", "", "",
             json.dumps(details)[:900]],
            value_input_option="RAW")
    except Exception as e:  # noqa: BLE001
        _out(f"WARN: _Run_Log append failed (outcome unaffected): {e}")


# --------------------------------------------------------------------------- #
# Modes                                                                       #
# --------------------------------------------------------------------------- #
def run_backup(dry_run: bool) -> int:
    sheet_id = _env("DEFAULT_SPREADSHEET_ID") or _env("SPREADSHEET_ID")
    folder_id = _env("TFB_BACKUP_FOLDER_ID")
    if not sheet_id:
        _out("ERROR: DEFAULT_SPREADSHEET_ID not set")
        return 2
    creds = _credentials()
    sess = _session(creds)

    t0 = time.monotonic()
    data = export_xlsx(sess, sheet_id)
    sha = sha256_hex(data)[:12]
    min_bytes = _env_int("TFB_BACKUP_MIN_BYTES", 1_000_000)
    _out(f"export: {len(data):,} bytes in {time.monotonic()-t0:.1f}s "
         f"sha256={sha}")
    if len(data) < min_bytes:
        msg = (f"[BACKUP v{__version__}] FAIL export too small: "
               f"{len(data)} < {min_bytes} bytes")
        _out("ERROR: " + msg)
        append_run_log("FAIL", msg, {"bytes": len(data), "sha": sha})
        return 3

    if dry_run:
        _out("dry-run — export verified, nothing uploaded")
        return 0
    if not folder_id:
        email = getattr(creds, "_tfb_client_email", "?")
        msg = (f"[BACKUP v{__version__}] FAIL TFB_BACKUP_FOLDER_ID not set — "
               f"create a Drive folder, share EDITOR with {email}, add the "
               f"repo Variable")
        _out("ERROR: " + msg)
        append_run_log("FAIL", msg, {"setup": "folder_id_missing"})
        return 2

    today = _now_riyadh().date()
    try:
        up = drive_upload_resumable(sess, folder_id, daily_name(today), data)
    except Exception as e:  # noqa: BLE001
        email = getattr(creds, "_tfb_client_email", "?")
        msg = (f"[BACKUP v{__version__}] FAIL upload: {e} — is the folder "
               f"shared EDITOR with {email}?")
        _out("ERROR: " + msg)
        append_run_log("FAIL", msg, {"stage": "upload"})
        return 3
    _out(f"uploaded {up.get('name')} id={up.get('id')}")

    listing = drive_list(sess, folder_id)
    names = [f.get("name", "") for f in listing]

    monthly_done = ""
    if monthly_needed(today, names):
        mn = monthly_name(today)
        try:
            drive_copy(sess, up["id"], folder_id, mn)
            monthly_done = mn
            names.append(mn)
            _out(f"monthly copy created: {mn}")
        except Exception as e:  # noqa: BLE001
            _out(f"WARN: monthly copy failed ({e}) — daily backup stands")

    keep_d = _env_int("TFB_BACKUP_KEEP_DAILY", 30)
    keep_m = _env_int("TFB_BACKUP_KEEP_MONTHLY", 12)
    _keep, delete = plan_pruning(names, keep_d, keep_m)
    by_name = {f.get("name"): f.get("id") for f in listing}
    deleted = 0
    for n in delete:
        fid = by_name.get(n)
        if not fid:
            continue
        try:
            drive_delete(sess, fid)
            deleted += 1
        except Exception as e:  # noqa: BLE001
            _out(f"WARN: prune failed for {n}: {e}")

    d_count = sum(1 for n in names if _DAILY_RE.match(n)) - deleted \
        if deleted else sum(1 for n in names if _DAILY_RE.match(n))
    m_count = sum(1 for n in names if _MONTHLY_RE.match(n))
    msg = (f"[BACKUP v{__version__}] name={daily_name(today)} "
           f"bytes={len(data)} sha={sha} monthly={monthly_done or '-'} "
           f"pruned={deleted} daily~{min(d_count, keep_d)}/{keep_d} "
           f"monthly~{min(m_count, keep_m)}/{keep_m}")
    _out(msg)
    append_run_log("OK", msg, {"bytes": len(data), "sha": sha,
                               "monthly": monthly_done, "pruned": deleted})
    return 0


def run_restore_test() -> int:
    """Acceptance criterion (Register §8): a backup never restored is a hope.
    Downloads the newest DAILY_, opens it, asserts the core tabs, logs."""
    folder_id = _env("TFB_BACKUP_FOLDER_ID")
    if not folder_id:
        _out("ERROR: TFB_BACKUP_FOLDER_ID not set")
        return 2
    sess = _session(_credentials())
    listing = drive_list(sess, folder_id)
    dailies = sorted((f for f in listing
                      if _DAILY_RE.match(f.get("name", ""))),
                     key=lambda f: f.get("name", ""), reverse=True)
    if not dailies:
        msg = f"[RESTORE-TEST v{__version__}] FAIL no DAILY_ backup found"
        _out("ERROR: " + msg)
        append_run_log("FAIL", msg, {})
        return 3
    target = dailies[0]
    t0 = time.monotonic()
    data = drive_download(sess, target["id"])
    from openpyxl import load_workbook  # CI-only import
    wb = load_workbook(io.BytesIO(data), read_only=True)
    tabs = set(wb.sheetnames)
    missing = sorted(_CORE_TABS - tabs)
    dur = time.monotonic() - t0
    if missing:
        msg = (f"[RESTORE-TEST v{__version__}] FAIL {target['name']} "
               f"missing tabs: {missing}")
        _out("ERROR: " + msg)
        append_run_log("FAIL", msg, {"tabs": len(tabs), "missing": missing})
        return 3
    msg = (f"[RESTORE-TEST v{__version__}] PASS {target['name']} "
           f"bytes={len(data)} tabs={len(tabs)} core_tabs=all "
           f"open_time={dur:.1f}s")
    _out(msg)
    append_run_log("OK", msg, {"bytes": len(data), "tabs": len(tabs),
                               "seconds": round(dur, 1)})
    return 0


# --------------------------------------------------------------------------- #
# Offline selftest                                                            #
# --------------------------------------------------------------------------- #
def _selftest() -> int:
    checks: List[Tuple[str, bool]] = []
    d = _dt.date(2026, 7, 22)
    checks.append(("names deterministic",
                   daily_name(d) == "DAILY_TFB_Workbook_2026-07-22.xlsx"
                   and monthly_name(d) == "MONTHLY_TFB_Workbook_2026-07.xlsx"))
    names = ([daily_name(_dt.date(2026, 7, 1) + _dt.timedelta(days=i))
              for i in range(35)]
             + [f"MONTHLY_TFB_Workbook_2025-{m:02d}.xlsx"
                for m in range(1, 13)]
             + ["MONTHLY_TFB_Workbook_2026-06.xlsx",
                "MONTHLY_TFB_Workbook_2026-07.xlsx",
                "operator_notes.txt"])
    keep, delete = plan_pruning(names, 30, 12)
    checks.append(("prune: 30 newest dailies kept",
                   sum(1 for n in keep if _DAILY_RE.match(n)) == 30
                   and daily_name(_dt.date(2026, 8, 4)) in keep
                   and daily_name(_dt.date(2026, 7, 1)) in delete))
    checks.append(("prune: 12 newest monthlies kept, oldest dropped",
                   sum(1 for n in keep if _MONTHLY_RE.match(n)) == 12
                   and "MONTHLY_TFB_Workbook_2025-01.xlsx" in delete
                   and "MONTHLY_TFB_Workbook_2025-02.xlsx" in delete
                   and "MONTHLY_TFB_Workbook_2026-07.xlsx" in keep))
    checks.append(("prune: foreign files untouched",
                   "operator_notes.txt" not in keep + delete))
    checks.append(("monthly_needed: due once per month",
                   monthly_needed(d, ["MONTHLY_TFB_Workbook_2026-06.xlsx"])
                   and not monthly_needed(
                       d, ["MONTHLY_TFB_Workbook_2026-07.xlsx"])))
    h1, h2 = sha256_hex(b"tfb"), sha256_hex(b"tfb")
    checks.append(("sha stable 64-hex", h1 == h2 and len(h1) == 64))
    checks.append(("env int guard", _env_int("TFB_NOPE_X", 30) == 30))
    passed = sum(1 for _, ok in checks if ok)
    for name, ok in checks:
        print(("PASS " if ok else "FAIL ") + name)
    print(f"[workbook_backup v{__version__}] SELFTEST {passed}/{len(checks)}")
    return 0 if passed == len(checks) else 1


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="TFB nightly workbook backup")
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--dry-run", action="store_true")
    g.add_argument("--restore-test", action="store_true")
    g.add_argument("--selftest", action="store_true")
    args = ap.parse_args(argv)
    if args.selftest:
        return _selftest()
    try:
        if args.restore_test:
            return run_restore_test()
        return run_backup(dry_run=bool(args.dry_run))
    except Exception as e:  # noqa: BLE001
        msg = f"[BACKUP v{__version__}] FAIL unhandled: {type(e).__name__}: {e}"
        _out("ERROR: " + msg)
        append_run_log("FAIL", msg, {"stage": "unhandled"})
        return 4


if __name__ == "__main__":
    raise SystemExit(main())
