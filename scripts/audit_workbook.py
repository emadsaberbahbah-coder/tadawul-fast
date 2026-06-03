#!/usr/bin/env python3
"""
scripts/audit_workbook.py
================================================================================
TADAWUL FAST BRIDGE — OFFLINE WORKBOOK AUDITOR (v1.2.0)
================================================================================
Static / file-based / network-free / deterministic / engine-contract-aware

HOW TO USE IT (the easy way — nothing to set up)
------------------------------------------------
  1. Open your Google Sheet -> File -> Download -> Microsoft Excel (.xlsx).
  2. Run:  python audit_workbook.py
That's it. The script installs what it needs, finds the file you just
downloaded (it checks Downloads, Desktop, Documents, and the current folder),
audits it, prints a plain-English report, and saves three report files into a
"tfb_audit" folder. No flags, no pip, no paths to remember.

If auto-detection ever picks the wrong file, point it at one directly:
  python audit_workbook.py -w /path/to/your_export.xlsx

Why this exists
---------------
`audit_data_quality.py` (the live auditor) talks to the FastAPI backend and
Google Sheets. It cannot inspect an exported `.xlsx` snapshot offline. This
module is the *export-file counterpart*: it ingests a downloaded workbook and
audits every recognised data page against the engine's 115-key canonical
contract — required fields, magic-number sentinels, fraction/percent scale
drift, 52W coherence, subunit (GBp/ZAC/ILA) market-cap normalisation, structural
fallback clusters, the v5.74+ investability gate, and sheet hygiene.

It is ADDITIVE: it does not import, replace, or alter the live auditor or any
core engine module. Stdlib + openpyxl (auto-installed on first run). Import-safe.

Design choices that prevent false positives
--------------------------------------------
- Field scale is per-field, not per-name. `Percent Change`, `Dividend Yield`,
  margins, ROI, volatility, `Upside %`, `Unrealized P/L %` are stored as DECIMAL
  FRACTIONS (0.0374 == 3.74%). `52W Position %` and every *Score* column are
  0–100. `Overall Penalty Factor` is [0,1]. These classes are checked with
  different bounds — matching the engine's anti-percent guards.
- `Max Drawdown 1Y` uses an inconsistent sign across pages in real exports, so
  it is range-checked per row but only flagged when its sign convention is
  *mixed within a page* (a comparability issue, not a per-row error).
- Gate status/action coherence uses a data-validated contradiction deny-list,
  so the intended WATCHLIST+DO_NOT_INVEST pairing is never flagged.
- "Real row" == non-empty Symbol. Thousands of trailing pre-formatted blank
  rows are counted as HYGIENE, never as data defects.
- Field resolution is by HEADER NAME, so pages of differing width all resolve.

Outputs (written automatically into ./tfb_audit/)
-------------------------------------------------
  audit_summary.txt   plain-English report, complete findings list
  audit_report.json   machine-readable
  audit_findings.csv  one row per finding (opens in Excel)

Exit codes (stable)
-------------------
  0 = clean (no WARN/HIGH/CRITICAL)
  1 = at least one WARN
  2 = at least one HIGH
  3 = at least one CRITICAL  (also used for "no workbook found")
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

def _ensure_openpyxl():
    """Import openpyxl; if it isn't installed, install it automatically (one time)
    so the user never has to run pip themselves."""
    try:
        from openpyxl import load_workbook as _lw
        return _lw
    except Exception:
        pass
    print("[setup] openpyxl not found — installing it for you (one-time)…")
    import subprocess
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"]
        )
    except Exception as exc:  # pragma: no cover
        print("FATAL: could not auto-install openpyxl. Run this once manually:\n"
              "    %s -m pip install openpyxl\nDetail: %s" % (sys.executable, exc),
              file=sys.stderr)
        sys.exit(3)
    from openpyxl import load_workbook as _lw  # retry after install
    print("[setup] openpyxl installed.")
    return _lw


load_workbook = _ensure_openpyxl()

__version__ = "1.2.0"

# =============================================================================
# Severity model
# =============================================================================
SEV_INFO = "INFO"
SEV_WARN = "WARN"
SEV_HIGH = "HIGH"
SEV_CRIT = "CRITICAL"

_SEV_RANK = {SEV_INFO: 0, SEV_WARN: 1, SEV_HIGH: 2, SEV_CRIT: 3}
_EXIT_FOR_RANK = {0: 0, 1: 1, 2: 2, 3: 3}


# =============================================================================
# Canonical contract — engine 115-key instrument schema
# (header order is authoritative; keys are metadata for the report)
# =============================================================================
CANONICAL_HEADERS: Tuple[str, ...] = (
    "Symbol", "Name", "Asset Class", "Exchange", "Currency", "Country", "Sector",
    "Industry", "Current Price", "Previous Close", "Open", "Day High", "Day Low",
    "52W High", "52W Low", "Price Change", "Percent Change", "52W Position %",
    "Volume", "Avg Volume 10D", "Avg Volume 30D", "Market Cap", "Float Shares",
    "Beta (5Y)", "P/E (TTM)", "P/E (Forward)", "EPS (TTM)", "Dividend Yield",
    "Payout Ratio", "Revenue (TTM)", "Revenue Growth YoY", "Gross Margin",
    "Operating Margin", "Profit Margin", "Debt/Equity", "Free Cash Flow (TTM)",
    "RSI (14)", "Volatility 30D", "Volatility 90D", "Max Drawdown 1Y",
    "VaR 95% (1D)", "Sharpe (1Y)", "Risk Score", "Risk Bucket", "P/B", "P/S",
    "EV/EBITDA", "PEG", "Intrinsic Value", "Upside %", "Valuation Score",
    "Forecast Price 1M", "Forecast Price 3M", "Forecast Price 12M",
    "Expected ROI 1M", "Expected ROI 3M", "Expected ROI 12M", "Forecast Confidence",
    "Confidence Score", "Confidence Bucket", "Value Score", "Quality Score",
    "Momentum Score", "Growth Score", "Overall Score", "Opportunity Score",
    "Rank (Overall)", "Fundamental View", "Technical View", "Risk View",
    "Value View", "Recommendation", "Recommendation Reason", "Horizon Days",
    "Invest Period Label", "Position Qty", "Avg Cost", "Position Cost",
    "Position Value", "Unrealized P/L", "Unrealized P/L %", "Data Provider",
    "Last Updated (UTC)", "Last Updated (Riyadh)", "Warnings", "Sector-Adj Score",
    "Conviction Score", "Top Factors", "Top Risks", "Position Size Hint",
    "Recommendation Detail", "Reco Priority", "Provider Rating",
    "Recommendation Source", "Priority Band", "Scoring Reco Source",
    "Scoring Schema Version", "Scoring Errors", "Opportunity Source",
    "Overall Score (Raw)", "Overall Penalty Factor", "Candle Pattern",
    "Candle Signal", "Candle Strength", "Candle Confidence", "Recent Patterns (5D)",
    "Forecast Source", "Data Quality Score", "Forecast Reliability Score",
    "Provider/Engine Conflict", "Conflict Type", "Final Decision Basis",
    "Investability Status", "Final Action", "Block Reason",
)
CANONICAL_LAST = "Block Reason"  # marks the end of the canonical data block

HEADER_TO_KEY: Dict[str, str] = {
    "Symbol": "symbol", "Name": "name", "Asset Class": "asset_class",
    "Exchange": "exchange", "Currency": "currency", "Country": "country",
    "Sector": "sector", "Industry": "industry", "Current Price": "current_price",
    "Previous Close": "previous_close", "Open": "open_price", "Day High": "day_high",
    "Day Low": "day_low", "52W High": "week_52_high", "52W Low": "week_52_low",
    "Price Change": "price_change", "Percent Change": "percent_change",
    "52W Position %": "week_52_position_pct", "Volume": "volume",
    "Avg Volume 10D": "avg_volume_10d", "Avg Volume 30D": "avg_volume_30d",
    "Market Cap": "market_cap", "Float Shares": "float_shares", "Beta (5Y)": "beta_5y",
    "P/E (TTM)": "pe_ttm", "P/E (Forward)": "pe_forward", "EPS (TTM)": "eps_ttm",
    "Dividend Yield": "dividend_yield", "Payout Ratio": "payout_ratio",
    "Revenue (TTM)": "revenue_ttm", "Revenue Growth YoY": "revenue_growth_yoy",
    "Gross Margin": "gross_margin", "Operating Margin": "operating_margin",
    "Profit Margin": "profit_margin", "Debt/Equity": "debt_to_equity",
    "Free Cash Flow (TTM)": "free_cash_flow_ttm", "RSI (14)": "rsi_14",
    "Volatility 30D": "volatility_30d", "Volatility 90D": "volatility_90d",
    "Max Drawdown 1Y": "max_drawdown_1y", "VaR 95% (1D)": "var_95_1d",
    "Sharpe (1Y)": "sharpe_1y", "Risk Score": "risk_score", "Risk Bucket": "risk_bucket",
    "P/B": "pb_ratio", "P/S": "ps_ratio", "EV/EBITDA": "ev_ebitda", "PEG": "peg_ratio",
    "Intrinsic Value": "intrinsic_value", "Upside %": "upside_pct",
    "Valuation Score": "valuation_score", "Forecast Price 1M": "forecast_price_1m",
    "Forecast Price 3M": "forecast_price_3m", "Forecast Price 12M": "forecast_price_12m",
    "Expected ROI 1M": "expected_roi_1m", "Expected ROI 3M": "expected_roi_3m",
    "Expected ROI 12M": "expected_roi_12m", "Forecast Confidence": "forecast_confidence",
    "Confidence Score": "confidence_score", "Confidence Bucket": "confidence_bucket",
    "Value Score": "value_score", "Quality Score": "quality_score",
    "Momentum Score": "momentum_score", "Growth Score": "growth_score",
    "Overall Score": "overall_score", "Opportunity Score": "opportunity_score",
    "Rank (Overall)": "rank_overall", "Fundamental View": "fundamental_view",
    "Technical View": "technical_view", "Risk View": "risk_view",
    "Value View": "value_view", "Recommendation": "recommendation",
    "Recommendation Reason": "recommendation_reason", "Horizon Days": "horizon_days",
    "Invest Period Label": "invest_period_label", "Position Qty": "position_qty",
    "Avg Cost": "avg_cost", "Position Cost": "position_cost",
    "Position Value": "position_value", "Unrealized P/L": "unrealized_pl",
    "Unrealized P/L %": "unrealized_pl_pct", "Data Provider": "data_provider",
    "Last Updated (UTC)": "last_updated_utc", "Last Updated (Riyadh)": "last_updated_riyadh",
    "Warnings": "warnings", "Sector-Adj Score": "sector_adj_score",
    "Conviction Score": "conviction_score", "Top Factors": "top_factors",
    "Top Risks": "top_risks", "Position Size Hint": "position_size_hint",
    "Recommendation Detail": "recommendation_detail", "Reco Priority": "reco_priority",
    "Provider Rating": "provider_rating", "Recommendation Source": "recommendation_source",
    "Priority Band": "priority_band", "Scoring Reco Source": "scoring_reco_source",
    "Scoring Schema Version": "scoring_schema_version", "Scoring Errors": "scoring_errors",
    "Opportunity Source": "opportunity_source", "Overall Score (Raw)": "overall_score_raw",
    "Overall Penalty Factor": "overall_penalty_factor", "Candle Pattern": "candle_pattern",
    "Candle Signal": "candle_signal", "Candle Strength": "candle_strength",
    "Candle Confidence": "candle_confidence", "Recent Patterns (5D)": "recent_patterns_5d",
    "Forecast Source": "forecast_source", "Data Quality Score": "data_quality_score",
    "Forecast Reliability Score": "forecast_reliability_score",
    "Provider/Engine Conflict": "provider_engine_conflict", "Conflict Type": "conflict_type",
    "Final Decision Basis": "final_decision_basis",
    "Investability Status": "investability_status", "Final Action": "final_action",
    "Block Reason": "block_reason",
}

# -----------------------------------------------------------------------------
# Field-class registries (the heart of false-positive avoidance)
# -----------------------------------------------------------------------------
IDENTITY_FIELDS = {"Symbol", "Name", "Asset Class", "Exchange", "Currency",
                   "Country", "Sector", "Industry"}

# All stored as DECIMAL FRACTIONS (0.0374 == 3.74%); may be negative.
# Split by how tightly the magnitude is bounded in legitimate data, so we don't
# false-flag genuinely high-growth / high-volatility / deep-value names.
#
# HARD: a value past ~1.5 (150%) is almost certainly percent-scale leakage.
FRACTION_HARD_FIELDS = {
    "Percent Change", "Dividend Yield", "Gross Margin",
    "Operating Margin", "Profit Margin",
}
FRACTION_HARD_LIMIT = 1.5
# SOFT: legitimately large for volatile/forecast-driven names; only extreme
# magnitudes (>5x == 500%) suggest contamination.
FRACTION_SOFT_FIELDS = {
    "Payout Ratio", "Revenue Growth YoY", "Volatility 30D", "Volatility 90D",
    "Expected ROI 1M", "Expected ROI 3M", "Expected ROI 12M",
    "Upside %", "Unrealized P/L %",
}
FRACTION_SOFT_LIMIT = 5.0
# Max Drawdown is bounded in [-1, 0]; VaR is a small (usually negative) fraction.
DRAWDOWN_FIELD = "Max Drawdown 1Y"
VAR_FIELD = "VaR 95% (1D)"

# Stored on a 0–100 scale.
SCORE_0_100_FIELDS = {
    "Risk Score", "Valuation Score", "Confidence Score", "Value Score",
    "Quality Score", "Momentum Score", "Growth Score", "Overall Score",
    "Opportunity Score", "Sector-Adj Score", "Conviction Score",
    "Data Quality Score", "Forecast Reliability Score", "Overall Score (Raw)",
}
SCORE_TOL = 0.5  # allow tiny float overshoot past 0/100

# 0–100 but semantically a position, not a score.
POSITION_0_100_FIELDS = {"52W Position %"}

# Strictly [0,1].
PENALTY_0_1_FIELDS = {"Overall Penalty Factor"}

# Price-like; subject to magic-sentinel + 52W coherence checks.
PRICE_FAMILY = {
    "Current Price", "Previous Close", "Open", "Day High", "Day Low",
    "52W High", "52W Low", "Forecast Price 1M", "Forecast Price 3M",
    "Forecast Price 12M", "Intrinsic Value",
}
# Non-negative counts.
COUNT_FAMILY = {"Volume", "Avg Volume 10D", "Avg Volume 30D", "Float Shares"}

# Columns scanned for "identical value repeated across N rows" fallback clusters.
CLUSTER_FIELDS = {
    "Value Score", "Quality Score", "Momentum Score", "Growth Score",
    "Overall Score", "Opportunity Score", "Valuation Score", "Confidence Score",
    "Sector-Adj Score", "Risk Score", "Beta (5Y)", "RSI (14)",
    "Gross Margin", "Operating Margin", "Profit Margin",
}

# Investability gate
GATE_STATUS_FIELD = "Investability Status"
GATE_ACTION_FIELD = "Final Action"
GATE_BLOCK_FIELD = "Block Reason"
GATE_VALID_STATUS = {"INVESTABLE", "WATCHLIST", "BLOCKED"}
GATE_VALID_ACTION = {"INVEST", "WATCH", "DO_NOT_INVEST", "REDUCE", "HOLD"}
# (status, action) pairs that are genuine contradictions. Derived empirically
# from the live export's joint distribution rather than an a-priori allow-list:
# the only pairs that actually occur are WATCHLIST+DO_NOT_INVEST (the intended
# "watch it, don't buy yet" pairing — NOT a conflict), WATCHLIST+WATCH,
# INVESTABLE+INVEST and BLOCKED+DO_NOT_INVEST, all coherent. We therefore flag
# only the combinations that would be self-contradictory if they appeared.
GATE_CONTRADICTIONS: set = {
    ("INVESTABLE", "DO_NOT_INVEST"),
    ("INVESTABLE", "REDUCE"),
    ("BLOCKED", "INVEST"),
    ("BLOCKED", "WATCH"),
    ("WATCHLIST", "INVEST"),
}

# Currencies quoted in a 1/100 subunit (pence / agorot / cents) — market-cap
# normalisation (÷100) is required or the cap inflates ~100×.
SUBUNIT_CURRENCIES = {"GBp", "GBX", "ZAC", "ILA", "ZAr", "KWf"}

# Magic sentinels that leak in as fallbacks.
EXACT_MAGIC = {999999.9999, 999999.0, -999999.9999}

# Pages we know how to audit as instrument tables.
KNOWN_DATA_PAGES = (
    "Market_Leaders", "Global_Markets", "Commodities_FX",
    "Mutual_Funds", "My_Portfolio", "Top_10_Investments",
)
# Soft expectations on row magnitude (informational only).
PAGE_ROW_EXPECT: Dict[str, Tuple[int, int, str]] = {
    "Top_10_Investments": (8, 12, "Top-10 page should hold ~10 ranked rows"),
    "My_Portfolio": (1, 200, "Personal portfolio is expected to be small"),
    "Commodities_FX": (5, 400, "Commodities/FX universe is a curated shortlist"),
}


# =============================================================================
# Finding model
# =============================================================================
@dataclass
class Finding:
    page: str
    severity: str
    code: str
    field: str
    symbol: str
    sheet_row: Optional[int]
    message: str
    value: Any = None

    def as_dict(self) -> Dict[str, Any]:
        return {
            "page": self.page,
            "severity": self.severity,
            "code": self.code,
            "field": self.field,
            "symbol": self.symbol,
            "sheet_row": self.sheet_row,
            "message": self.message,
            "value": _jsonable(self.value),
        }


# =============================================================================
# Helpers
# =============================================================================
def _jsonable(v: Any) -> Any:
    if v is None or isinstance(v, (bool, int, float, str)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return str(v)
        return v
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v)


def _is_num(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool) \
        and not (isinstance(v, float) and (math.isnan(v) or math.isinf(v)))


def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _is_magic(v: Any) -> bool:
    """999999.9999-family + any mantissa of >=5 consecutive 9s."""
    if not _is_num(v):
        return False
    f = float(v)
    if f in EXACT_MAGIC:
        return True
    mant = ("%.4f" % abs(f)).replace(".", "").rstrip("0")
    return len(mant) >= 5 and set(mant) == {"9"}


def _norm_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# =============================================================================
# Page reader
# =============================================================================
@dataclass
class PageData:
    name: str
    headers: List[str]
    canonical_width: int
    rows: List[Tuple]           # raw row tuples
    col_index: Dict[str, int]   # header -> first column index
    real_row_numbers: List[int]  # 1-based sheet row number of each REAL row
    real_rows: List[Tuple]       # rows whose Symbol is non-blank
    total_rows: int
    empty_rows: int


def read_page(wb, name: str) -> PageData:
    ws = wb[name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = list(next(rows_iter))
    except StopIteration:
        header_row = []
    headers = [(_norm_str(h) if h is not None else None) for h in header_row]

    if CANONICAL_LAST in headers:
        canonical_width = headers.index(CANONICAL_LAST) + 1
    else:
        canonical_width = len([h for h in headers if h])

    col_index: Dict[str, int] = {}
    for i, h in enumerate(headers[:canonical_width]):
        if h and h not in col_index:
            col_index[h] = i

    si = col_index.get("Symbol", 0)

    rows: List[Tuple] = []
    real_rows: List[Tuple] = []
    real_row_numbers: List[int] = []
    total = 0
    empty = 0
    for offset, r in enumerate(rows_iter, start=2):
        rows.append(r)
        total += 1
        sym = r[si] if si < len(r) else None
        if _is_blank(sym):
            empty += 1
        else:
            real_rows.append(r)
            real_row_numbers.append(offset)

    return PageData(
        name=name, headers=headers, canonical_width=canonical_width, rows=rows,
        col_index=col_index, real_row_numbers=real_row_numbers, real_rows=real_rows,
        total_rows=total, empty_rows=empty,
    )


def _cell(row: Tuple, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# =============================================================================
# Individual checks — each appends Findings and returns nothing
# =============================================================================
def check_header_integrity(pd: PageData, out: List[Finding]) -> None:
    # Canonical block must match the engine contract in order, where present.
    missing = [h for h in CANONICAL_HEADERS[:pd.canonical_width] if h not in pd.col_index]
    if missing and pd.name != "Top_10_Investments":
        out.append(Finding(pd.name, SEV_HIGH, "C1_HEADER_MISSING", "<schema>", "",
                            None, "Canonical headers absent from sheet: "
                            + ", ".join(missing[:8]) + ("…" if len(missing) > 8 else "")))

    # Contamination: repeated/garbage headers BEYOND the canonical block.
    tail = pd.headers[pd.canonical_width:]
    tail_nonempty = [h for h in tail if h]
    repeats = [h for h, c in Counter(tail_nonempty).items() if c > 1]
    if repeats:
        worst = Counter(tail_nonempty).most_common(1)[0]
        out.append(Finding(
            pd.name, SEV_WARN, "C1_HEADER_CONTAMINATION", "<header-row>", "", 1,
            "Header row carries a repeated label past the canonical block "
            "(status-writeback overrun?): '%s' x%d across the tail columns."
            % (worst[0], worst[1]), value=worst[0]))


def check_required_and_hygiene(pd: PageData, out: List[Finding]) -> None:
    if pd.empty_rows > 0:
        pct = (pd.empty_rows / pd.total_rows * 100.0) if pd.total_rows else 0.0
        sev = SEV_WARN if pct >= 50.0 else SEV_INFO
        out.append(Finding(
            pd.name, sev, "C2_EMPTY_ROW_BLOAT", "<sheet>", "", None,
            "%d of %d rows are empty (%.1f%%). Pre-formatted/uncleared rows bloat "
            "the export and slow reads; a hard clear before each refresh is advised."
            % (pd.empty_rows, pd.total_rows, pct), value=pd.empty_rows))

    pi = pd.col_index.get("Current Price")
    for rownum, row in zip(pd.real_row_numbers, pd.real_rows):
        sym = _norm_str(_cell(row, pd.col_index.get("Symbol")))
        if pi is not None:
            price = _cell(row, pi)
            if _is_blank(price):
                out.append(Finding(pd.name, SEV_CRIT, "C3_MISSING_PRICE",
                                    "Current Price", sym, rownum,
                                    "Required field Current Price is empty."))
            elif _is_num(price) and float(price) == 0.0:
                out.append(Finding(pd.name, SEV_HIGH, "C3_ZERO_PRICE",
                                    "Current Price", sym, rownum,
                                    "Current Price is exactly 0.0 (missing-as-zero?).",
                                    value=0.0))


def check_magic_sentinels(pd: PageData, out: List[Finding]) -> None:
    for rownum, row in zip(pd.real_row_numbers, pd.real_rows):
        sym = _norm_str(_cell(row, pd.col_index.get("Symbol")))
        for h in PRICE_FAMILY:
            idx = pd.col_index.get(h)
            if idx is None:
                continue
            v = _cell(row, idx)
            if _is_magic(v):
                out.append(Finding(pd.name, SEV_HIGH, "C4_MAGIC_SENTINEL", h, sym,
                                   rownum, "Magic sentinel leaked into a price field "
                                   "(%s). Distorts 52W range / position." % h, value=v))


def check_scale_drift(pd: PageData, out: List[Finding]) -> None:
    _dd_signs: List[int] = []  # collected drawdown signs for a page-level consistency check
    for rownum, row in zip(pd.real_row_numbers, pd.real_rows):
        sym = _norm_str(_cell(row, pd.col_index.get("Symbol")))

        # HARD fraction fields: a value past 1.5 (150%) is almost certainly
        # a ×100 percent-scale leak.
        for h in FRACTION_HARD_FIELDS:
            idx = pd.col_index.get(h)
            v = _cell(row, idx) if idx is not None else None
            if _is_num(v) and abs(float(v)) > FRACTION_HARD_LIMIT:
                out.append(Finding(pd.name, SEV_HIGH, "C5_FRACTION_DRIFT", h, sym,
                                   rownum, "%s = %s exceeds |%.1f|; expected a decimal "
                                   "fraction (×100 contamination?)."
                                   % (h, v, FRACTION_HARD_LIMIT), value=v))

        # SOFT fraction fields: legitimately large for volatile / forecast-driven
        # names; only extreme magnitudes (>5× == 500%) suggest contamination.
        for h in FRACTION_SOFT_FIELDS:
            idx = pd.col_index.get(h)
            v = _cell(row, idx) if idx is not None else None
            if _is_num(v) and abs(float(v)) > FRACTION_SOFT_LIMIT:
                out.append(Finding(pd.name, SEV_WARN, "C5_FRACTION_EXTREME", h, sym,
                                   rownum, "%s = %s exceeds |%.1f| (500%%); plausible but "
                                   "extreme — verify it is not percent-scale leakage."
                                   % (h, v, FRACTION_SOFT_LIMIT), value=v))

        # Max Drawdown: the engine's sign convention varies by page/provider
        # (some store the magnitude positive, some negative). A value is only a
        # genuine scale error if |v| exceeds 1.0 (a drawdown cannot exceed 100%).
        # Sign *consistency* is checked once per page, below the row loop.
        idx = pd.col_index.get(DRAWDOWN_FIELD)
        v = _cell(row, idx) if idx is not None else None
        if _is_num(v):
            fv = float(v)
            if abs(fv) > 1.0 + 1e-9:
                out.append(Finding(pd.name, SEV_HIGH, "C5_DRAWDOWN_OUT_OF_RANGE",
                                   DRAWDOWN_FIELD, sym, rownum,
                                   "%s = %s implies >100%% drawdown." % (DRAWDOWN_FIELD, v),
                                   value=v))
            if fv > 1e-6:
                _dd_signs.append(1)
            elif fv < -1e-6:
                _dd_signs.append(-1)

        # VaR (1D) is a small fraction; |v| > 1.0 (100% daily) is implausible.
        idx = pd.col_index.get(VAR_FIELD)
        v = _cell(row, idx) if idx is not None else None
        if _is_num(v) and abs(float(v)) > 1.0:
            out.append(Finding(pd.name, SEV_WARN, "C5_VAR_SUSPICIOUS",
                               VAR_FIELD, sym, rownum,
                               "%s = %s implies >100%% 1-day loss; verify scale."
                               % (VAR_FIELD, v), value=v))

        for h in SCORE_0_100_FIELDS:
            idx = pd.col_index.get(h)
            v = _cell(row, idx) if idx is not None else None
            if _is_num(v) and (float(v) < -SCORE_TOL or float(v) > 100.0 + SCORE_TOL):
                out.append(Finding(pd.name, SEV_HIGH, "C5_SCORE_OUT_OF_RANGE", h, sym,
                                   rownum, "%s = %s outside 0–100." % (h, v), value=v))

        for h in POSITION_0_100_FIELDS:
            idx = pd.col_index.get(h)
            v = _cell(row, idx) if idx is not None else None
            if _is_num(v) and (float(v) < -SCORE_TOL or float(v) > 100.0 + SCORE_TOL):
                out.append(Finding(pd.name, SEV_WARN, "C5_POSITION_OUT_OF_RANGE", h, sym,
                                   rownum, "%s = %s outside 0–100." % (h, v), value=v))

        for h in PENALTY_0_1_FIELDS:
            idx = pd.col_index.get(h)
            v = _cell(row, idx) if idx is not None else None
            if _is_num(v) and (float(v) < -1e-9 or float(v) > 1.0 + 1e-9):
                out.append(Finding(pd.name, SEV_HIGH, "C5_PENALTY_OUT_OF_RANGE", h, sym,
                                   rownum, "%s = %s outside [0,1]." % (h, v), value=v))

        for h in COUNT_FAMILY:
            idx = pd.col_index.get(h)
            v = _cell(row, idx) if idx is not None else None
            if _is_num(v) and float(v) < 0:
                out.append(Finding(pd.name, SEV_WARN, "C5_NEGATIVE_COUNT", h, sym,
                                   rownum, "%s = %s is negative." % (h, v), value=v))

    # Page-level: Max Drawdown should use one sign convention within a page.
    # A meaningful mix of positive and negative values means the sign is
    # inconsistent (a downstream comparability hazard), even though each value
    # is individually in-range.
    if _dd_signs:
        npos = _dd_signs.count(1)
        nneg = _dd_signs.count(-1)
        minor = min(npos, nneg)
        if minor > 0:
            out.append(Finding(pd.name, SEV_WARN, "C5_DRAWDOWN_SIGN_MIXED",
                               DRAWDOWN_FIELD, "", None,
                               "%s mixes sign conventions on this page (%d positive, "
                               "%d negative); values are in-range but not comparable "
                               "across rows without normalisation."
                               % (DRAWDOWN_FIELD, npos, nneg), value=minor))


def check_52w_coherence(pd: PageData, out: List[Finding]) -> None:
    pi = pd.col_index.get("Current Price")
    hi_i = pd.col_index.get("52W High")
    lo_i = pd.col_index.get("52W Low")
    pos_i = pd.col_index.get("52W Position %")
    if pi is None or hi_i is None or lo_i is None:
        return
    for rownum, row in zip(pd.real_row_numbers, pd.real_rows):
        sym = _norm_str(_cell(row, pd.col_index.get("Symbol")))
        p, hi, lo = _cell(row, pi), _cell(row, hi_i), _cell(row, lo_i)
        if not (_is_num(p) and _is_num(hi) and _is_num(lo)):
            continue
        if _is_magic(hi) or _is_magic(lo):
            continue  # already reported by C4
        p, hi, lo = float(p), float(hi), float(lo)
        if hi <= 0 or lo <= 0:
            continue
        if lo > hi:
            out.append(Finding(pd.name, SEV_HIGH, "C6_RANGE_INVERTED", "52W Low/High",
                               sym, rownum, "52W Low %.4g > 52W High %.4g." % (lo, hi)))
            continue
        if p > hi * 1.002:
            out.append(Finding(pd.name, SEV_WARN, "C6_PRICE_ABOVE_52W_HIGH", "52W High",
                               sym, rownum, "Price %.4g exceeds stale 52W High %.4g."
                               % (p, hi), value=hi))
        elif p < lo * 0.998:
            out.append(Finding(pd.name, SEV_WARN, "C6_PRICE_BELOW_52W_LOW", "52W Low",
                               sym, rownum, "Price %.4g below stale 52W Low %.4g."
                               % (p, lo), value=lo))
        # Position% consistency
        if pos_i is not None:
            pos = _cell(row, pos_i)
            if _is_num(pos) and hi > lo:
                expected = (p - lo) / (hi - lo) * 100.0
                if abs(float(pos) - expected) > 5.0 and 0 <= expected <= 100:
                    out.append(Finding(pd.name, SEV_WARN, "C6_POSITION_MISMATCH",
                                       "52W Position %", sym, rownum,
                                       "52W Position %% %.1f disagrees with computed %.1f."
                                       % (float(pos), expected), value=float(pos)))


def check_market_cap_subunit(pd: PageData, out: List[Finding]) -> None:
    ci = pd.col_index.get("Currency")
    mci = pd.col_index.get("Market Cap")
    pi = pd.col_index.get("Current Price")
    fi = pd.col_index.get("Float Shares")
    if ci is None or mci is None or pi is None or fi is None:
        return
    for rownum, row in zip(pd.real_row_numbers, pd.real_rows):
        ccy = _norm_str(_cell(row, ci))
        if ccy not in SUBUNIT_CURRENCIES:
            continue
        mc, price, fs = _cell(row, mci), _cell(row, pi), _cell(row, fi)
        if not (_is_num(mc) and _is_num(price) and _is_num(fs)):
            continue
        mc, price, fs = float(mc), float(price), float(fs)
        if mc <= 0 or price <= 0 or fs <= 0:
            continue
        implied = mc / price
        ratio = implied / fs  # ~1 => mcap shares price's subunit => inflated ~100x
        sym = _norm_str(_cell(row, pd.col_index.get("Symbol")))
        if 0.5 <= ratio <= 2.0:
            out.append(Finding(pd.name, SEV_HIGH, "C7_SUBUNIT_MCAP", "Market Cap", sym,
                               rownum, "Currency %s (1/100 subunit): implied/float shares "
                               "ratio %.3f ~ 1 → Market Cap appears in the subunit and is "
                               "~100x a major-currency cap (÷100 normalisation missing)."
                               % (ccy, ratio), value=mc))


def check_fallback_clusters(pd: PageData, out: List[Finding], min_cluster: int) -> None:
    n_real = len(pd.real_rows)
    if n_real < max(min_cluster, 10):
        return
    for h in CLUSTER_FIELDS:
        idx = pd.col_index.get(h)
        if idx is None:
            continue
        vals = [round(float(_cell(r, idx)), 4) for r in pd.real_rows if _is_num(_cell(r, idx))]
        if len(vals) < min_cluster:
            continue
        common_val, count = Counter(vals).most_common(1)[0]
        share = count / len(vals)
        if count >= min_cluster and share >= 0.10:
            out.append(Finding(pd.name, SEV_WARN, "C8_FALLBACK_CLUSTER", h, "", None,
                               "Identical value %.4g repeated in %d/%d populated rows "
                               "(%.0f%%) — likely a default/fallback injection."
                               % (common_val, count, len(vals), share * 100.0),
                               value=common_val))


def check_investability_gate(pd: PageData, out: List[Finding], summary: Dict[str, Any]) -> None:
    si = pd.col_index.get(GATE_STATUS_FIELD)
    ai = pd.col_index.get(GATE_ACTION_FIELD)
    bi = pd.col_index.get(GATE_BLOCK_FIELD)
    if si is None:
        summary["gate_present"] = False
        out.append(Finding(pd.name, SEV_WARN, "C9_GATE_ABSENT",
                           GATE_STATUS_FIELD, "", None,
                           "Investability gate columns are absent on this page; rows are "
                           "ungated (v5.74+ gate not applied here)."))
        return
    summary["gate_present"] = True
    status_dist: Counter = Counter()
    populated = 0
    for rownum, row in zip(pd.real_row_numbers, pd.real_rows):
        sym = _norm_str(_cell(row, pd.col_index.get("Symbol")))
        status = _norm_str(_cell(row, si)).upper()
        action = _norm_str(_cell(row, ai)).upper() if ai is not None else ""
        block = _norm_str(_cell(row, bi)) if bi is not None else ""
        if not status:
            out.append(Finding(pd.name, SEV_HIGH, "C9_GATE_UNPOPULATED",
                               GATE_STATUS_FIELD, sym, rownum,
                               "Real row has an empty Investability Status."))
            continue
        populated += 1
        status_dist[status] += 1
        if status not in GATE_VALID_STATUS:
            out.append(Finding(pd.name, SEV_HIGH, "C9_GATE_BAD_STATUS",
                               GATE_STATUS_FIELD, sym, rownum,
                               "Unknown Investability Status '%s'." % status, value=status))
            continue
        if action and action not in GATE_VALID_ACTION:
            out.append(Finding(pd.name, SEV_WARN, "C9_GATE_BAD_ACTION",
                               GATE_ACTION_FIELD, sym, rownum,
                               "Unknown Final Action '%s'." % action, value=action))
        if action and (status, action) in GATE_CONTRADICTIONS:
            out.append(Finding(pd.name, SEV_HIGH, "C9_GATE_STATUS_ACTION_CONFLICT",
                               GATE_ACTION_FIELD, sym, rownum,
                               "Status %s contradicts Final Action %s."
                               % (status, action), value=action))
        if status == "BLOCKED" and not block:
            out.append(Finding(pd.name, SEV_HIGH, "C9_BLOCKED_NO_REASON",
                               GATE_BLOCK_FIELD, sym, rownum,
                               "Status BLOCKED but Block Reason is empty."))
        if status == "INVESTABLE" and block:
            out.append(Finding(pd.name, SEV_WARN, "C9_INVESTABLE_HAS_BLOCK",
                               GATE_BLOCK_FIELD, sym, rownum,
                               "Status INVESTABLE yet a Block Reason is present.",
                               value=block))
    summary["gate_populated"] = populated
    summary["gate_status_distribution"] = dict(status_dist)


def check_scoring_population(pd: PageData, out: List[Finding], summary: Dict[str, Any]) -> None:
    idx = pd.col_index.get("Overall Score")
    if idx is None:
        return
    populated = sum(1 for r in pd.real_rows if _is_num(_cell(r, idx)))
    n = len(pd.real_rows)
    summary["overall_score_populated"] = populated
    if n and populated < n:
        gap = n - populated
        pct = gap / n * 100.0
        sev = SEV_WARN if pct >= 10.0 else SEV_INFO
        out.append(Finding(pd.name, sev, "C10_SCORING_GAP", "Overall Score", "", None,
                           "%d of %d real rows (%.1f%%) lack an Overall Score."
                           % (gap, n, pct), value=gap))


def check_duplicate_symbols(pd: PageData, out: List[Finding]) -> None:
    si = pd.col_index.get("Symbol")
    if si is None:
        return
    seen: Dict[str, int] = {}
    for rownum, row in zip(pd.real_row_numbers, pd.real_rows):
        sym = _norm_str(_cell(row, si))
        if not sym:
            continue
        if sym in seen:
            out.append(Finding(pd.name, SEV_HIGH, "C11_DUPLICATE_SYMBOL", "Symbol", sym,
                               rownum, "Duplicate symbol (first seen at sheet row %d)."
                               % seen[sym], value=sym))
        else:
            seen[sym] = rownum


def check_page_shape(pd: PageData, out: List[Finding]) -> None:
    exp = PAGE_ROW_EXPECT.get(pd.name)
    if not exp:
        return
    lo, hi, why = exp
    n = len(pd.real_rows)
    if n < lo:
        out.append(Finding(pd.name, SEV_WARN, "C12_ROW_SHORTFALL", "<page>", "", None,
                           "%d real rows < expected minimum %d. %s" % (n, lo, why), value=n))
    elif n > hi:
        out.append(Finding(pd.name, SEV_INFO, "C12_ROW_EXCESS", "<page>", "", None,
                           "%d real rows > expected maximum %d. %s" % (n, hi, why), value=n))


# =============================================================================
# Page audit orchestration
# =============================================================================
def audit_page(wb, name: str, min_cluster: int) -> Tuple[Dict[str, Any], List[Finding]]:
    pd = read_page(wb, name)
    findings: List[Finding] = []
    summary: Dict[str, Any] = {
        "page": name,
        "total_rows": pd.total_rows,
        "real_rows": len(pd.real_rows),
        "empty_rows": pd.empty_rows,
        "canonical_width": pd.canonical_width,
        "headers_present": len(pd.col_index),
    }

    check_header_integrity(pd, findings)
    check_required_and_hygiene(pd, findings)
    check_magic_sentinels(pd, findings)
    check_scale_drift(pd, findings)
    check_52w_coherence(pd, findings)
    check_market_cap_subunit(pd, findings)
    check_fallback_clusters(pd, findings, min_cluster)
    check_investability_gate(pd, findings, summary)
    check_scoring_population(pd, findings, summary)
    check_duplicate_symbols(pd, findings)
    check_page_shape(pd, findings)

    sev_counts = Counter(f.severity for f in findings)
    summary["severity_counts"] = {s: sev_counts.get(s, 0)
                                  for s in (SEV_INFO, SEV_WARN, SEV_HIGH, SEV_CRIT)}
    summary["finding_codes"] = dict(Counter(f.code for f in findings))
    return summary, findings


# =============================================================================
# Report assembly + outputs
# =============================================================================
def build_report(workbook: str, page_summaries: List[Dict[str, Any]],
                 all_findings: List[Finding]) -> Dict[str, Any]:
    global_sev = Counter(f.severity for f in all_findings)
    global_codes = Counter(f.code for f in all_findings)
    max_rank = max([_SEV_RANK[f.severity] for f in all_findings], default=0)
    return {
        "auditor": "TFB Offline Workbook Auditor",
        "auditor_version": __version__,
        "generated_utc": _utc_now_iso(),
        "workbook": workbook,
        "canonical_contract_keys": len(CANONICAL_HEADERS),
        "pages_audited": [s["page"] for s in page_summaries],
        "global_severity_counts": {s: global_sev.get(s, 0)
                                   for s in (SEV_INFO, SEV_WARN, SEV_HIGH, SEV_CRIT)},
        "global_finding_codes": dict(global_codes),
        "max_severity": [k for k, v in _SEV_RANK.items() if v == max_rank][0],
        "page_summaries": page_summaries,
        "findings": [f.as_dict() for f in all_findings],
    }


def write_json(report: Dict[str, Any], path: str) -> None:
    Path(path).write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")


def write_csv(all_findings: List[Finding], path: str) -> None:
    cols = ["page", "severity", "code", "field", "symbol", "sheet_row", "value", "message"]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for f in all_findings:
            d = f.as_dict()
            w.writerow({c: d.get(c, "") for c in cols})


def _report_lines(report: Dict[str, Any], include_info: bool) -> List[str]:
    """Build the human-readable report as a list of lines (used for both the
    console and the saved .txt summary, so they never drift apart)."""
    g = report["global_severity_counts"]
    L: List[str] = []
    L.append("=" * 78)
    L.append("TFB OFFLINE WORKBOOK AUDITOR  v%s" % report["auditor_version"])
    L.append("Workbook : %s" % report["workbook"])
    L.append("Generated: %s" % report["generated_utc"])
    L.append("=" * 78)
    L.append("PER-PAGE SUMMARY")
    L.append("-" * 78)
    L.append("%-20s %6s %6s %6s | %4s %4s %4s %4s" % (
        "Page", "rows", "real", "empty", "INFO", "WARN", "HIGH", "CRIT"))
    for s in report["page_summaries"]:
        sc = s["severity_counts"]
        gate = ""
        if s.get("gate_present"):
            gate = "  gate:%d %s" % (s.get("gate_populated", 0),
                                     s.get("gate_status_distribution", {}))
        elif s.get("gate_present") is False:
            gate = "  gate:absent"
        L.append("%-20s %6d %6d %6d | %4d %4d %4d %4d%s" % (
            s["page"], s["total_rows"], s["real_rows"], s["empty_rows"],
            sc[SEV_INFO], sc[SEV_WARN], sc[SEV_HIGH], sc[SEV_CRIT], gate))
    L.append("-" * 78)
    L.append("GLOBAL: INFO=%d  WARN=%d  HIGH=%d  CRITICAL=%d   (max severity: %s)" % (
        g[SEV_INFO], g[SEV_WARN], g[SEV_HIGH], g[SEV_CRIT], report["max_severity"]))
    L.append("Finding codes: " + json.dumps(report["global_finding_codes"]))
    L.append("=" * 78)

    shown = [f for f in report["findings"]
             if include_info or f["severity"] != SEV_INFO]
    shown.sort(key=lambda f: (-_SEV_RANK[f["severity"]], f["page"], f["code"]))
    if shown:
        L.append("FINDINGS (most severe first; %d shown)" % len(shown))
        L.append("-" * 78)
        for f in shown:
            loc = ("row %s" % f["sheet_row"]) if f["sheet_row"] else "-"
            sym = (" [%s]" % f["symbol"]) if f["symbol"] else ""
            L.append("[%-8s] %-22s %-20s %s%s — %s" % (
                f["severity"], f["code"], f["page"], loc, sym, f["message"]))
    else:
        L.append("No findings at or above WARN. The workbook looks clean.")
    L.append("=" * 78)
    return L


def print_console(report: Dict[str, Any], include_info: bool) -> None:
    # Console caps the findings list; the saved .txt holds the complete list.
    g = report["global_severity_counts"]
    lines = _report_lines(report, include_info)
    # Find where the FINDINGS section starts so we can cap just that part.
    out: List[str] = []
    in_findings = False
    findings_emitted = 0
    for ln in lines:
        if ln.startswith("FINDINGS ("):
            in_findings = True
        if in_findings and ln.startswith("["):
            findings_emitted += 1
            if findings_emitted == 61:
                out.append("… more findings — see the saved report files.")
            if findings_emitted >= 61:
                continue
        out.append(ln)
    print("\n".join(out))


def write_summary_txt(report: Dict[str, Any], path: str, include_info: bool) -> None:
    Path(path).write_text("\n".join(_report_lines(report, include_info)) + "\n",
                          encoding="utf-8")


# =============================================================================
# CLI
# =============================================================================
def discover_workbook(explicit: Optional[str]) -> Optional[Path]:
    """Find the workbook to audit.

    If the user passed a path, use it. Otherwise look in the usual places a
    downloaded sheet lands and pick the most recently modified .xlsx, so the
    user can just run the script with no arguments after downloading.
    """
    if explicit:
        p = Path(explicit).expanduser()
        return p if p.exists() else None

    home = Path.home()
    search_dirs = [
        Path.cwd(),
        home / "Downloads",
        home / "Desktop",
        home / "Documents",
        Path(__file__).resolve().parent,
    ]
    seen: set = set()
    candidates: List[Path] = []
    for d in search_dirs:
        try:
            if not d.exists() or d in seen:
                continue
            seen.add(d)
            for f in d.glob("*.xlsx"):
                # Skip Excel lock/temp files like "~$Book.xlsx".
                if f.name.startswith("~$"):
                    continue
                candidates.append(f)
        except Exception:
            continue
    if not candidates:
        return None
    # Newest first.
    candidates.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return candidates[0]


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="TFB offline workbook auditor. Just run it after downloading "
                    "the Google Sheet as Excel — it finds the file and writes the "
                    "report automatically. All flags are optional.")
    p.add_argument("--workbook", "-w", default=None,
                   help="Path to the exported .xlsx (optional; auto-detected if omitted).")
    p.add_argument("--pages", nargs="*", default=None,
                   help="Subset of data pages to audit (default: all recognised).")
    p.add_argument("--out-dir", default=None,
                   help="Folder for the report files (default: ./tfb_audit next to where you run it).")
    p.add_argument("--json-out", default=None, help="Override the JSON report path.")
    p.add_argument("--csv-out", default=None, help="Override the CSV findings path.")
    p.add_argument("--min-cluster", type=int, default=25,
                   help="Min identical-value count to flag a fallback cluster (default 25).")
    p.add_argument("--include-info", action="store_true",
                   help="Show INFO findings in the console too.")
    p.add_argument("--quiet", action="store_true", help="Suppress the console report.")
    return p.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)

    # 1) Find the workbook (explicit path, else auto-detect).
    wb_file = discover_workbook(args.workbook)
    if wb_file is None:
        if args.workbook:
            print("FATAL: workbook not found: %s" % args.workbook, file=sys.stderr)
        else:
            print("=" * 78, file=sys.stderr)
            print("Couldn't find a workbook to audit.", file=sys.stderr)
            print("Do this:", file=sys.stderr)
            print("  1. Open your Google Sheet.", file=sys.stderr)
            print("  2. File -> Download -> Microsoft Excel (.xlsx).", file=sys.stderr)
            print("  3. Run this script again (it checks Downloads/Desktop/this folder).", file=sys.stderr)
            print("Or point it straight at the file:", file=sys.stderr)
            print("     python audit_workbook.py -w /path/to/your_export.xlsx", file=sys.stderr)
            print("=" * 78, file=sys.stderr)
        return 3
    wb_path = str(wb_file)
    if not args.quiet:
        print("Auditing workbook: %s" % wb_path)

    # 2) Decide where the report files go (auto-created).
    out_dir = Path(args.out_dir).expanduser() if args.out_dir else (Path.cwd() / "tfb_audit")
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = args.json_out or str(out_dir / "audit_report.json")
    csv_path = args.csv_out or str(out_dir / "audit_findings.csv")
    txt_path = str(out_dir / "audit_summary.txt")

    # 3) Load and select pages.
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    sheetnames = set(wb.sheetnames)
    if args.pages:
        targets = [p for p in args.pages if p in sheetnames]
        skipped = [p for p in args.pages if p not in sheetnames]
        for s in skipped:
            print("WARN: requested page not in workbook, skipping: %s" % s, file=sys.stderr)
    else:
        targets = [p for p in KNOWN_DATA_PAGES if p in sheetnames]
    if not targets:
        print("FATAL: no recognised data pages found to audit. Sheets present: %s"
              % sorted(sheetnames), file=sys.stderr)
        return 3

    # 4) Audit.
    page_summaries: List[Dict[str, Any]] = []
    all_findings: List[Finding] = []
    for name in targets:
        summary, findings = audit_page(wb, name, args.min_cluster)
        page_summaries.append(summary)
        all_findings.extend(findings)
    report = build_report(wb_path, page_summaries, all_findings)

    # 5) Always write all three report files.
    write_json(report, json_path)
    write_csv(all_findings, csv_path)
    write_summary_txt(report, txt_path, include_info=True)

    # 6) Console report + where-to-find-it footer.
    if not args.quiet:
        print_console(report, include_info=args.include_info)
        print("Report files written to: %s" % out_dir)
        print("  - audit_summary.txt   (plain-English, full findings list)")
        print("  - audit_report.json   (machine-readable)")
        print("  - audit_findings.csv  (one row per finding, open in Excel)")
        verdict = report["max_severity"]
        if verdict == SEV_CRIT:
            print("VERDICT: CRITICAL — there are rows that can't be acted on. See the summary.")
        elif verdict == SEV_HIGH:
            print("VERDICT: HIGH — some values look wrong. See the summary.")
        elif verdict == SEV_WARN:
            print("VERDICT: WARN — minor/structural issues only.")
        else:
            print("VERDICT: CLEAN — nothing at or above WARN.")

    max_rank = _SEV_RANK[report["max_severity"]]
    return _EXIT_FOR_RANK.get(max_rank, 0)


if __name__ == "__main__":
    sys.exit(main())
