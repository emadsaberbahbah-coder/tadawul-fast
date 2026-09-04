#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tfb_acceptance.py — v1.0.0 (2026-08-31, 10-day finalization program, Day 6)
================================================================================
WHY THIS EXISTS
    "Done" in this program is a printed table of MEASUREMENTS, never an opinion.
    Every morning audit of the last month re-derived the same numbers by hand
    (row counts, blank Target Price, format seam, overdue cohorts, false greens,
    capacity key, calibration key, decision-feed truth). This script computes
    them from an export directory (the browser TSVs + optional xlsx) or from the
    LIVE workbook (CI, service account) and prints the Day-10 definition of done
    with the measured value next to each criterion. Both AIs run it blind on the
    same artifact (W7 constraint b: re-execute every numeric claim).

WHAT IT MEASURES (each row = criterion, measured value, verdict, evidence)
    D10-1 board fills          Top_10 KPI Passed / Selected / funding alerts
    D10-2 evidence clock       TFB Grid Capacity key + newest backend _Run_Log row age
    D10-3 learning loop        Performance_Log matured/active/overdue/duplicates,
                               TFB Calibration key
    D10-4 feed truthful        TFB Decision Feed key; SUCCESS-over-PARTIAL stamps
    D10-5 one writer per page  display-glyph share per page; Open outside [L,H]
    D10-6 daily brief          NA from the workbook (workflow state); reported as NA
    D10-7 S-1 clock            S1_Gate scored days / verdict
    G1..G4 gates               rows, dup/blank symbols, freshness, provider_target
                               share, Target Price share, false greens (INVEST with
                               fetch_failed or a non-ticker symbol)
    A1..A4 decision integrity  (v1.0.6, 2026-09-04 six-gate audit classes)
                               A1 board seats on Derayah-tradable venues (symbol
                                  suffix in core.compliance_gate.DERAYAH_MARKETS)
                               A2 USD FX peg: Top_10 audit + Portfolio_Decision
                                  Price SAR / Price within [3.74, 3.77]
                               A3 equity identity collisions: futures / crypto /
                                  contract-month names on Global_Markets /
                                  Market_Leaders equity rows (KE.US -> wheat,
                                  LINK.US -> Chainlink, NG.US -> natural gas)
                               A4 Portfolio_Decision TRIM/EXIT rows carry Δ Shares

VERDICT VOCABULARY   PASS / WARN / FAIL / NA. Exit code 0 unless --strict and
any FAIL (the CI workflow runs non-strict: an instrument, not a blocker).

USAGE
    python scripts/tfb_acceptance.py --export-dir /path/to/exports [--xlsx file]
    python scripts/tfb_acceptance.py --live [--sheet-id ID]     # CI, creds in env
    python scripts/tfb_acceptance.py --selftest                # offline proof
    Options: --json artifacts/acceptance.json  --strict  --board-min 5

ENV (live mode)   DEFAULT_SPREADSHEET_ID | SPREADSHEET_ID | TARGET_SHEET_ID
                  GOOGLE_SHEETS_CREDENTIALS(_B64) | GOOGLE_APPLICATION_CREDENTIALS
NO WRITES, EVER.
"""
from __future__ import annotations

import argparse
import base64
import csv
import datetime as _dt
import json
import os
import hashlib
import re
import subprocess
import sys
from typing import Any, Dict, List, Optional, Tuple

VERSION = "1.0.6"
PAGES = ("Market_Leaders", "Global_Markets", "Commodities_FX", "Mutual_Funds")
_TICKER_RE = re.compile(r"^[A-Z0-9^][A-Z0-9.\-=^&/]{0,23}$")
_NUM_RE = re.compile(r"^[+-]?\d*\.?\d+(?:[eE][+-]?\d+)?$")
RIYADH = _dt.timezone(_dt.timedelta(hours=3))

# v1.0.6 — decision-integrity constants. The venue set mirrors
# core.compliance_gate.DERAYAH_MARKETS (loaded live when importable; the frozen
# copy below is the offline fallback and is reported as such in evidence).
_DERAYAH_SUFFIXES_FROZEN = frozenset((
    "US", "SR", "T", "HK", "L", "PA", "AS", "BR", "DE", "MI", "MC", "LS", "VI",
    "SW", "TO", "AX", "OL", "SI", "MX"))
_FX_PEG_BAND = (3.74, 3.77)          # SAR/USD official 3.75; spot 3.7500..3.7550
_FX_PEG_MIN_ROWS = 5
# names that can never belong to an equity row: futures, contract months, crypto pairs
_IDENTITY_COLLISION_RE = re.compile(
    r"\bfutures\b"          # plural only: "Future FinTech Group" is a company
    r"|\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s-]?\d{2}(?:\d{2})?\s*$"
    r"|,\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*-\d{4}\b"
    r"|\susd\s*$",
    re.I)
_NON_EQUITY_SYMBOL_RE = re.compile(r"(=F|=X|-USD|-USDT|\^)", re.I)


# --------------------------------------------------------------------------- #
# small helpers (pure)                                                        #
# --------------------------------------------------------------------------- #
def _now_riyadh() -> _dt.datetime:
    return _dt.datetime.now(RIYADH)


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _to_float(v: Any) -> Optional[float]:
    t = _s(v).replace(",", "").replace("%", "").replace("\u25b2", "").replace("\u25bc", "").strip()
    if not t or not _NUM_RE.match(t):
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _is_glyph(v: Any) -> bool:
    t = _s(v)
    return t.startswith(("\u25b2", "\u25bc")) or t.endswith("%")


def _parse_dt(v: Any) -> Optional[_dt.datetime]:
    t = _s(v)
    if not t:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d %H:%M:%S%z",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y"):
        try:
            d = _dt.datetime.strptime(t[:32] if "T" in t else t, fmt)
            return d if d.tzinfo else d.replace(tzinfo=RIYADH)
        except ValueError:
            continue
    try:
        d = _dt.datetime.fromisoformat(t[:32])
        return d if d.tzinfo else d.replace(tzinfo=RIYADH)
    except Exception:
        return None


def _sha256_file(path: str) -> str:
    h = hashlib.sha256()
    try:
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1 << 20), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""


def _mtime_iso(path: str) -> str:
    try:
        return _dt.datetime.fromtimestamp(os.path.getmtime(path), tz=_dt.timezone.utc).isoformat()
    except Exception:
        return ""


def _table(rows: List[List[str]]) -> List[Dict[str, str]]:
    """header row + data rows -> list of dicts (header-keyed, blank-safe)."""
    if not rows:
        return []
    hdr = [_s(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or not any(_s(c) for c in r):
            continue
        d = {}
        for i, h in enumerate(hdr):
            if h:
                d[h] = r[i] if i < len(r) else ""
        out.append(d)
    return out


class Check:
    def __init__(self, cid: str, name: str, verdict: str, measured: Any, evidence: str = ""):
        self.cid, self.name, self.verdict, self.measured, self.evidence = cid, name, verdict, measured, evidence

    def row(self) -> Dict[str, Any]:
        return {"id": self.cid, "criterion": self.name, "verdict": self.verdict,
                "measured": self.measured, "evidence": self.evidence}


# --------------------------------------------------------------------------- #
# data sources                                                                #
# --------------------------------------------------------------------------- #
class Source:
    """Uniform access: tab(name) -> list of rows (list of str). Missing -> [].
    v1.0.2: every read problem is RECORDED in .errors (fail-closed: a check on
    an unreadable tab reports NA with the reason; --strict fails on errors)."""
    errors: List[str] = []
    inputs: List[Dict[str, Any]] = []

    def tab(self, name: str) -> List[List[str]]:  # pragma: no cover - abstract
        raise NotImplementedError

    def runlog_tail(self, n: int = 400) -> List[List[str]]:
        rows = self.tab("_Run_Log")
        return rows[-n:] if rows else []


class ExportSource(Source):
    def __init__(self, export_dir: str, xlsx: str = ""):
        self.dir = export_dir
        self.xlsx = xlsx
        self._cache: Dict[str, List[List[str]]] = {}
        self._files: Dict[str, str] = {}
        self.errors = []
        self.inputs = []
        cands: Dict[str, List[str]] = {}
        for fn in sorted(os.listdir(export_dir)):
            if fn.lower().endswith(".tsv") and "_-_" in fn:
                tab = fn.rsplit("_-_", 1)[1][:-4]
                # v1.0.1: browser download counters never identify a file
                # (project rule): "Global_Markets(6)" / "Global_Markets__6_" -> tab.
                tab = re.sub(r"(\s*\(\d+\)|__\d+_)\s*$", "", tab).strip()
                cands.setdefault(tab, []).append(os.path.join(export_dir, fn))
        # v1.0.2: deterministic choice — the NEWEST file by mtime (ties: name);
        # every candidate is recorded so the choice is auditable.
        for tab, paths in cands.items():
            chosen = sorted(paths, key=lambda q: (os.path.getmtime(q), q))[-1]
            self._files[tab] = chosen
            self.inputs.append({"tab": tab, "file": os.path.basename(chosen),
                                "sha256": _sha256_file(chosen), "mtime_utc": _mtime_iso(chosen),
                                "candidates": [os.path.basename(q) for q in paths]})
        if not xlsx:
            xl = sorted((fn for fn in os.listdir(export_dir) if fn.lower().endswith(".xlsx")),
                        key=lambda fn: (os.path.getmtime(os.path.join(export_dir, fn)), fn))
            if xl:
                self.xlsx = os.path.join(export_dir, xl[-1])
        if self.xlsx:
            if os.path.exists(self.xlsx):
                self.inputs.append({"tab": "*xlsx*", "file": os.path.basename(self.xlsx),
                                    "sha256": _sha256_file(self.xlsx), "mtime_utc": _mtime_iso(self.xlsx)})
            else:
                self.errors.append(f"xlsx not found: {self.xlsx}")

    def tab(self, name: str) -> List[List[str]]:
        if name in self._cache:
            return self._cache[name]
        rows: List[List[str]] = []
        if name in self._files:
            with open(self._files[name], encoding="utf-8", newline="") as fh:
                rows = [list(r) for r in csv.reader(fh, delimiter="\t", quoting=csv.QUOTE_NONE)]
        elif self.xlsx and os.path.exists(self.xlsx):
            try:
                from openpyxl import load_workbook  # optional dependency
            except Exception as exc:  # v1.0.2: never silent
                self.errors.append(f"openpyxl unavailable — xlsx tab {name} unread: {exc}")
                self._cache[name] = []
                return []
            try:
                wb = load_workbook(self.xlsx, read_only=True, data_only=True)
                if name in wb.sheetnames:
                    ws = wb[name]
                    rows = [[_s(c) for c in r] for r in ws.iter_rows(values_only=True)]
                else:
                    self.errors.append(f"tab {name} absent from xlsx and no TSV")
            except Exception as exc:
                self.errors.append(f"xlsx read failed for tab {name}: {exc}")
                rows = []
        else:
            self.errors.append(f"tab {name}: no TSV and no xlsx")
        self._cache[name] = rows
        return rows


class LiveSource(Source):
    def __init__(self, sheet_id: str):
        self.sheet_id = sheet_id
        self._cache: Dict[str, List[List[str]]] = {}
        self.errors = []
        self.inputs = [{"tab": "*live*", "workbook_id": sheet_id}]
        self.gc = self._client()
        self.book = self.gc.open_by_key(sheet_id)

    @staticmethod
    def _client():
        import gspread  # CI-only import
        from google.oauth2 import service_account
        scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
        path = _s(os.getenv("GOOGLE_APPLICATION_CREDENTIALS"))
        if path and os.path.exists(path):
            creds = service_account.Credentials.from_service_account_file(path, scopes=scopes)
        else:
            raw = (_s(os.getenv("GOOGLE_SHEETS_CREDENTIALS")) or _s(os.getenv("GOOGLE_SHEETS_CREDENTIALS_B64"))
                   or _s(os.getenv("GOOGLE_CREDENTIALS")))
            if not raw:
                raise RuntimeError("no Google credentials in env")
            s = raw
            if not s.startswith("{"):
                try:
                    dec = base64.b64decode(s).decode("utf-8", errors="replace").strip()
                    if dec.startswith("{"):
                        s = dec
                except Exception:
                    pass
            creds = service_account.Credentials.from_service_account_info(json.loads(s), scopes=scopes)
        return gspread.authorize(creds)

    _PAGE_COLS = ("Symbol", "Open", "Day High", "Day Low", "Expected ROI 12M", "Forecast Source",
                  "Target Price", "Investability Status", "Final Action", "Warnings", "Last Updated (UTC)")

    @staticmethod
    def _col_letter(n: int) -> str:
        out = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            out = chr(65 + r) + out
        return out

    def _rng(self, ws, r0: int, r1: int, cols: int) -> str:
        """v1.0.1: a range CLAMPED to the worksheet grid (values.get rejects a
        range beyond the grid with 'exceeds grid limits')."""
        rc = int(getattr(ws, "row_count", 0) or 0) or r1
        cc = int(getattr(ws, "col_count", 0) or 0) or cols
        r1 = max(r0, min(r1, rc))
        return f"A{r0}:{self._col_letter(max(1, min(cols, cc)))}{r1}"

    def _page_columns(self, ws) -> List[List[str]]:
        """v1.0.1: read only the columns the checks need (header-driven), via
        one batch_get — a 6,609 x 115 page is ~8 MB as a single values.get."""
        hdr = (ws.get(self._rng(ws, 1, 1, 200)) or [[]])[0]
        idx = {}
        for i, h in enumerate(hdr):
            if _s(h) in self._PAGE_COLS and _s(h) not in idx:
                idx[_s(h)] = i + 1
        if "Symbol" not in idx:
            return []
        rc = int(getattr(ws, "row_count", 0) or 0) or 2
        names = [n for n in self._PAGE_COLS if n in idx]
        ranges = [f"{self._col_letter(idx[n])}2:{self._col_letter(idx[n])}{rc}" for n in names]
        blocks = ws.batch_get(ranges) or []
        n_rows = max((len(b) for b in blocks), default=0)
        rows: List[List[str]] = [list(names)]
        for r in range(n_rows):
            row = []
            for b in blocks:
                cell = b[r][0] if r < len(b) and b[r] else ""
                row.append(_s(cell))
            if any(row):
                rows.append(row)
        return rows

    def tab(self, name: str) -> List[List[str]]:
        if name in self._cache:
            return self._cache[name]
        try:
            ws = self.book.worksheet(name)
            rc = int(getattr(ws, "row_count", 0) or 0)
            if name == "_Run_Log":
                rows = ws.get(self._rng(ws, max(1, rc - 400), rc, 10)) or []
            elif name == "_Status":
                rows = ws.get(self._rng(ws, 1, 60, 13)) or []
            elif name == "Top_10_Investments":
                rows = ws.get(self._rng(ws, 1, 60, 12)) or []
            elif name == "S1_Gate":
                rows = ws.get(self._rng(ws, 1, 20, 4)) or []
            elif name == "Performance_Log":
                rows = ws.get(self._rng(ws, 1, rc, 32)) or []
            elif name in PAGES:
                rows = self._page_columns(ws)
            else:
                rows = ws.get_all_values() or []
        except Exception as exc:  # recorded, never silent: the check reports NA
            self.errors.append(f"live tab {name} unreadable: {exc}")
            print(f"[acceptance] tab {name} unreadable: {exc}", file=sys.stderr)
            rows = []
        self._cache[name] = [[_s(c) for c in r] for r in rows]
        return self._cache[name]


# --------------------------------------------------------------------------- #
# measurements                                                                #
# --------------------------------------------------------------------------- #
def _status_globals(src: Source) -> Dict[str, str]:
    kv: Dict[str, str] = {}
    for r in src.tab("_Status"):
        if len(r) > 12 and _s(r[11]):
            kv[_s(r[11]).casefold()] = _s(r[12])
    return kv


def _status_stamps(src: Source) -> List[Dict[str, str]]:
    return _table(src.tab("_Status"))


def _top10(src: Source) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    rows = [[_s(c) for c in r] for r in src.tab("Top_10_Investments")]
    for i, r in enumerate(rows):
        if not r or not any(r):
            continue
        if r[0].startswith("Status:") and len(r) > 1:
            out.setdefault("last_run", r[1][:160])
        if "Deployable (SAR)" in r and i + 1 < len(rows):
            out["kpis"] = {h: v for h, v in zip(r, rows[i + 1]) if h}
        if r[0].upper().startswith("ALERTS"):
            kinds = []
            for rr in rows[i + 2:]:
                if not rr or not rr[0]:
                    break
                kinds.append((rr[0], rr[1] if len(rr) > 1 else ""))
            out["alerts"] = kinds
        if r[0].upper().startswith("DATA GAPS"):   # v1.0.3: first-fail distribution
            gates = []
            for rr in rows[i + 2:]:
                if not rr or not rr[0]:
                    break
                gates.append(f"{rr[0]}={rr[1] if len(rr) > 1 else ''}")
            out["gates"] = gates
    return out


def check_board(src: Source, board_min: int) -> List[Check]:
    t = _top10(src)
    k = t.get("kpis") or {}
    passed = _to_float(k.get("Passed"))
    sel = _s(k.get("Selected"))
    alerts = dict(t.get("alerts") or [])
    if not k:
        return [Check("D10-1", "board fills (Top_10 Passed >= %d)" % board_min, "NA", None, "Top_10 tab unreadable")]
    v = "PASS" if (passed or 0) >= board_min else ("WARN" if (passed or 0) > 0 else "FAIL")
    ev = f"Passed={passed:g} Selected={sel} | funding alerts: " + (", ".join(
        f"{a}={c}" for a, c in alerts.items() if a in ("rotation_proposal", "capital_call", "unfunded_candidates")) or "none")
    if t.get("gates"):
        ev += " | first-fail: " + ", ".join(t["gates"][:4])
    return [Check("D10-1", "board fills (Top_10 Passed >= %d)" % board_min, v, passed, ev + f" | {t.get('last_run', '')[:80]}")]


def check_evidence_clock(src: Source) -> List[Check]:
    kv = _status_globals(src)
    cap = kv.get("tfb grid capacity", "")
    out = []
    if not cap:
        out.append(Check("D10-2a", "capacity key present (TFB Grid Capacity)", "FAIL", None, "key absent from _Status L:M"))
    else:
        st = cap.split("|", 1)[0].strip()
        out.append(Check("D10-2a", "capacity state", "PASS" if st == "OK" else ("WARN" if st == "NEAR-LIMIT" else "FAIL"), st, cap[:120]))
    tail = src.runlog_tail(400)
    newest = None
    for r in tail:
        if len(r) > 2 and _s(r[2]) in ("run_dashboard_sync", "sync_hold", "track_performance"):
            d = _parse_dt(r[0])
            if d and (newest is None or d > newest):
                newest = d
    if newest is None:
        out.append(Check("D10-2b", "evidence clock (newest backend _Run_Log row)", "FAIL", None, "no backend row in the tail read"))
    else:
        age_h = (_now_riyadh() - newest.astimezone(RIYADH)).total_seconds() / 3600.0
        out.append(Check("D10-2b", "evidence clock (newest backend _Run_Log row age h)",
                         "PASS" if age_h <= 24 else "FAIL", round(age_h, 1), newest.isoformat()))
    return out


def check_learning(src: Source) -> List[Check]:
    rows = src.tab("Performance_Log")
    # find header row (contains 'Record ID')
    hi = next((i for i, r in enumerate(rows[:10]) if r and _s(r[0]) == "Record ID"), None)
    kv = _status_globals(src)
    cal = kv.get("tfb calibration", "")
    out = [Check("D10-3a", "calibration key present (TFB Calibration)", "PASS" if cal else "FAIL",
                 cal.split("|", 1)[0].strip() if cal else None, cal[:120])]
    if hi is None:
        out.append(Check("D10-3b", "cohorts (Performance_Log)", "NA", None, "Performance_Log unreadable"))
        return out
    recs = _table(rows[hi:])
    today = _now_riyadh().date()
    active = [r for r in recs if _s(r.get("Status")).lower() == "active"]
    matured = sum(1 for r in recs if _s(r.get("Status")).lower() == "matured")
    dup = sum(1 for r in recs if _s(r.get("Outcome")) == "DUPLICATE_KEY")
    overdue = 0
    for r in active:
        d = _parse_dt(r.get("Target Date (Riyadh)"))
        if d and d.date() < today:
            overdue += 1
    share = (overdue / len(active) * 100.0) if active else 0.0
    v = "PASS" if overdue == 0 else ("WARN" if share < 1.0 else "FAIL")
    out.append(Check("D10-3b", "learning loop (overdue active cohorts)", v, overdue,
                     f"records={len(recs)} active={len(active)} matured={matured} duplicate_key={dup} overdue_share={share:.1f}%"))
    return out


def check_feed_truth(src: Source) -> List[Check]:
    kv = _status_globals(src)
    feed = kv.get("tfb decision feed", "")
    head = feed.split("|", 1)[0].strip() if feed else ""
    out = [Check("D10-4a", "decision feed key present", "PASS" if feed else "FAIL", head or None, feed[:120])]
    # v1.0.2: presence is not permission — the STATE is its own criterion.
    if head.upper().startswith("EXECUTABLE"):
        out.append(Check("D10-4c", "decision feed state (EXECUTABLE = actionable)", "PASS", "EXECUTABLE", feed[:120]))
    elif head.upper().startswith("NOT_ACTIONABLE"):
        out.append(Check("D10-4c", "decision feed state (EXECUTABLE = actionable)", "WARN", head[:40],
                         "truthful non-actionable state — no investment/deployment permission"))
    else:
        out.append(Check("D10-4c", "decision feed state (EXECUTABLE = actionable)", "FAIL", head[:40] or None, "absent or unrecognised"))
    lies = []
    for st in _status_stamps(src):
        page, status, msg = _s(st.get("Page")), _s(st.get("Status")).upper(), _s(st.get("Message"))
        if page in PAGES and "data=PARTIAL" in msg and status == "SUCCESS":
            lies.append(page)
    out.append(Check("D10-4b", "no SUCCESS-over-PARTIAL page stamp (status-truth armed)",
                     "PASS" if not lies else "FAIL", len(lies), ", ".join(lies) or "none"))
    return out


def _stamp_guard(src: Source) -> Dict[str, Dict[str, Any]]:
    """v1.0.3: per page, the sync stamp's guard counters 'guard=pw:x/n,rb:y/n'."""
    out: Dict[str, Dict[str, Any]] = {}
    for st in _status_stamps(src):
        page = _s(st.get("Page"))
        m = re.search(r"guard=pw:(\d+)/(\d+),rb:(\d+)/(\d+)", _s(st.get("Message")))
        if page in PAGES and m:
            # v1.0.5: the sync v6.56.0 stamp carries "rb_tol=<tol>(+<delta>)" when its bounded
            # write-survival tolerance decided a DIVERGENT readback as COMPLETE.
            t = re.search(r"rb_tol=(\d+)\(\+(-?\d+)\)", _s(st.get("Message")))
            out[page] = {"pw": int(m.group(1)), "pw_n": int(m.group(2)), "rb": int(m.group(3)), "rb_n": int(m.group(4)),
                         "stamp": _s(st.get("Last Updated")),
                         "rb_tol": int(t.group(1)) if t else None}
    return out


def check_pages(src: Source) -> List[Check]:
    out = []
    now = _now_riyadh()
    guard = _stamp_guard(src)
    for p in PAGES:
        rows = _table(src.tab(p))
        n = len(rows)
        if n == 0:
            out.append(Check(f"G1-{p}", f"{p} readable", "NA", 0, "tab empty/unreadable"))
            continue
        syms = [_s(r.get("Symbol")) for r in rows]
        dup = len(syms) - len(set(s for s in syms if s))
        blank = sum(1 for s in syms if not s)
        fresh = 0
        for r in rows:
            d = _parse_dt(r.get("Last Updated (UTC)"))
            if d and (now - d.astimezone(RIYADH)).total_seconds() <= 24 * 3600:
                fresh += 1
        glyph = sum(1 for r in rows if _is_glyph(r.get("Expected ROI 12M")))
        pt_rows = [r for r in rows if _s(r.get("Forecast Source")) == "provider_target"]
        pt = len(pt_rows)
        tp = sum(1 for r in pt_rows if _s(r.get("Target Price")))  # v1.0.2: within the cohort
        oor = 0
        for r in rows:
            o, h, l = _to_float(r.get("Open")), _to_float(r.get("Day High")), _to_float(r.get("Day Low"))
            if o is not None and h is not None and l is not None and (o < l or o > h):
                oor += 1
        fg = 0
        for r in rows:
            if _s(r.get("Final Action")).upper() == "INVEST" or _s(r.get("Investability Status")).upper() == "INVESTABLE":
                if not _TICKER_RE.match(_s(r.get("Symbol")).upper()) or "fetch_failed" in _s(r.get("Warnings")).lower():
                    fg += 1
        fresh_share = fresh / n * 100.0
        g1_ok = dup == 0 and blank == 0 and fresh_share >= 95.0  # v1.0.2: stale is a FAIL
        out.append(Check(f"G1-{p}", f"{p} integrity (rows/dup/blank/fresh>=95%)", "PASS" if g1_ok else "FAIL",
                         n, f"dup_symbols={dup} blank_symbols={blank} fresh24h={fresh}/{n} ({fresh_share:.1f}%)"))
        gs = glyph / n * 100.0
        g = guard.get(p)
        if g and g["rb_n"] > 0:
            rb_share = g["rb"] / g["rb_n"] * 100.0
            # v1.0.5: measure WRITE SURVIVAL (rb - pw), not the raw readback count. The
            # prewrite-flagged rows are provider OHLC incoherence already present in the
            # payload (tol_excused etc.), not something the write did. PASS when nothing
            # survived the write or the sync's own bounded tolerance (v6.56.0 rb_tol)
            # decided the cohort COMPLETE; WARN when the residual is <= 1% of rows; FAIL above.
            delta = max(0, g["rb"] - g["pw"])
            d_share = delta / g["rb_n"] * 100.0
            tol = g.get("rb_tol")
            verdict = "PASS" if (delta == 0 or tol is not None) else ("WARN" if d_share <= 1.0 else "FAIL")
            out.append(Check(f"D10-5-{p}", f"{p} OHLC coherence after write (write-survival rows rb-pw)",
                             verdict, delta,
                             f"survived={delta} ({d_share:.2f}%) rb={g['rb']}/{g['rb_n']} ({rb_share:.1f}%) pw={g['pw']}/{g['pw_n']}"
                             + (f" rb_tol={tol}" if tol is not None else "")
                             + f" open_outside_range={oor} stamp={g['stamp'][:19]}"))
        else:
            out.append(Check(f"D10-5-{p}", f"{p} OHLC coherence after write (readback-flagged rows)", "NA", None,
                             f"no guard counters on the page stamp; open_outside_range={oor}"))
        # v1.0.3: display glyphs are a GAS number FORMAT (exports render display text; the pool is
        # read UNFORMATTED) — informational, never a writer verdict.
        out.append(Check(f"G4-{p}", f"{p} display number-format share % (info)", "WARN" if glyph else "PASS",
                         round(gs, 1), f"glyph_cells={glyph} open_outside_range={oor} (layout format, not a writer)"))
        out.append(Check(f"G2-{p}", f"{p} targets (Target Price nonblank within provider_target rows)", "PASS" if (pt == 0 or tp >= 0.8 * pt) else "FAIL",
                         tp, f"provider_target={pt} target_price_nonblank_in_cohort={tp}"))
        out.append(Check(f"G3-{p}", f"{p} false greens (INVEST with fetch_failed / non-ticker)", "PASS" if fg == 0 else "FAIL", fg, ""))
    return out


def check_s1(src: Source) -> List[Check]:
    rows = src.tab("S1_Gate")
    scored, verdict = None, ""
    for r in rows:
        line = " | ".join(_s(c) for c in r)
        m = re.search(r"(\d+)/28 scored", line)
        if m:
            scored = int(m.group(1))
        m2 = re.search(r"verdict:\s*([A-Z_]+)", line)
        if m2:
            verdict = m2.group(1)
    if scored is None:
        return [Check("D10-7", "S-1 evidence clock (scored days /28)", "NA", None, "S1_Gate unreadable")]
    return [Check("D10-7", "S-1 evidence clock (scored days /28)", "PASS" if scored >= 28 else "WARN", scored, f"verdict={verdict or '?'}")]


def _derayah_suffixes() -> Tuple[frozenset, str]:
    """v1.0.6: (suffix set, source). Live from core.compliance_gate when the
    repo is importable (CI runs from the repo root); frozen copy otherwise."""
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        root = os.path.dirname(here)
        if root not in sys.path:
            sys.path.insert(0, root)
        from core.compliance_gate import DERAYAH_MARKETS  # type: ignore
        live = frozenset(_s(k).lstrip(".").upper() for k in DERAYAH_MARKETS.keys() if _s(k))
        if live:
            return live, "compliance_gate"
    except Exception:
        pass
    return _DERAYAH_SUFFIXES_FROZEN, "frozen"


def _symbol_suffix(sym: str) -> str:
    s = _s(sym).upper()
    if not s:
        return ""
    return s.rsplit(".", 1)[1] if "." in s else "US"


def _top10_sections(src: Source) -> Dict[str, Tuple[List[str], List[List[str]]]]:
    """v1.0.6: named sections of the Top_10 export -> (header, body rows).
    board    = the header row that carries both Symbol and Stability
    qualified= header after a row starting with ALL QUALIFIED
    audit    = header after a row starting with CANDIDATES"""
    rows = [[_s(c) for c in r] for r in src.tab("Top_10_Investments")]
    out: Dict[str, Tuple[List[str], List[List[str]]]] = {}

    def body(i: int) -> List[List[str]]:
        b: List[List[str]] = []
        for rr in rows[i + 1:]:
            if not rr or not any(rr) or not (rr[0] or (len(rr) > 1 and rr[1])):
                break
            b.append(rr)
        return b
    for i, r in enumerate(rows):
        if "Symbol" not in r:
            continue
        prev = _s(rows[i - 1][0]).upper() if i > 0 and rows[i - 1] else ""
        if "Stability" in r and "board" not in out:
            out["board"] = (r, body(i))
        elif prev.startswith("ALL QUALIFIED") and "qualified" not in out:
            out["qualified"] = (r, body(i))
        elif prev.startswith("CANDIDATES") and "audit" not in out:
            out["audit"] = (r, body(i))
    return out


def _implied_fx(header: List[str], body: List[List[str]], ccy_col: str, px_col: str,
                pxsar_col: str) -> List[Tuple[str, float]]:
    ix = {h: i for i, h in enumerate(header)}
    if ccy_col not in ix or px_col not in ix or pxsar_col not in ix:
        return []
    out: List[Tuple[str, float]] = []
    for r in body:
        if len(r) <= max(ix[ccy_col], ix[px_col], ix[pxsar_col]):
            continue
        if _s(r[ix[ccy_col]]).upper() != "USD":
            continue
        px, ps = _to_float(r[ix[px_col]]), _to_float(r[ix[pxsar_col]])
        if px and ps and px > 0 and ps > 0:
            out.append((_s(r[ix.get("Symbol", 0)]), ps / px))
    return out


def check_decision_integrity(src: Source) -> List[Check]:
    """v1.0.6: A1..A4 — the four 2026-09-04 audit classes, measured daily."""
    out: List[Check] = []
    sec = _top10_sections(src)
    suffixes, sfx_src = _derayah_suffixes()
    lo, hi = _FX_PEG_BAND
    # A1 — board seats on tradable venues
    if "board" in sec:
        hdr, body = sec["board"]
        si = hdr.index("Symbol")
        seats = [_s(r[si]) for r in body if len(r) > si and _s(r[si])]
        bad = [f"{s}(.{_symbol_suffix(s)})" for s in seats if _symbol_suffix(s) not in suffixes]
        out.append(Check("A1", "board seats on Derayah-tradable venues (suffix in DERAYAH_MARKETS)",
                         "PASS" if (seats and not bad) else ("FAIL" if bad else "NA"), len(bad),
                         f"seats={len(seats)} untradable={' '.join(bad) or 'none'} | venue set: {sfx_src} ({len(suffixes)})"))
    else:
        out.append(Check("A1", "board seats on Derayah-tradable venues (suffix in DERAYAH_MARKETS)", "NA", None,
                         "Top_10 board section not found"))
    # A2 — USD FX peg (Top_10 audit, then Portfolio_Decision)
    if "audit" in sec:
        hdr, body = sec["audit"]
        vals = _implied_fx(hdr, body, "Ccy", "Price", "Price SAR")
        if len(vals) >= _FX_PEG_MIN_ROWS:
            xs = sorted(v for _, v in vals)
            med = xs[len(xs) // 2]
            outside = [(s, v) for s, v in vals if not (lo <= v <= hi)]
            share = len(outside) / float(len(vals))
            v = "FAIL" if not (lo <= med <= hi) else ("WARN" if share > 0.01 else "PASS")
            out.append(Check("A2a", f"USD FX peg — Top_10 audit Price SAR/Price within [{lo}, {hi}]", v, round(med, 4),
                             f"n={len(vals)} median={med:.4f} min={xs[0]:.4f} max={xs[-1]:.4f} outside={len(outside)} "
                             f"({share * 100:.1f}%) e.g. {' '.join(f'{s}={x:.4f}' for s, x in outside[:3]) or '-'}"))
        else:
            out.append(Check("A2a", f"USD FX peg — Top_10 audit Price SAR/Price within [{lo}, {hi}]", "NA", None,
                             f"USD rows with both prices: {len(vals)} (< {_FX_PEG_MIN_ROWS})"))
    else:
        out.append(Check("A2a", f"USD FX peg — Top_10 audit Price SAR/Price within [{lo}, {hi}]", "NA", None,
                         "Top_10 audit section not found"))
    pd_rows = [[_s(c) for c in r] for r in src.tab("Portfolio_Decision")]
    pd_hdr_i = next((i for i, r in enumerate(pd_rows) if "Action" in r and "Symbol" in r), None)
    if pd_hdr_i is None:
        out.append(Check("A2b", f"USD FX peg — Portfolio_Decision Price SAR/Price within [{lo}, {hi}]", "NA", None,
                         "Portfolio_Decision header not found"))
        out.append(Check("A4", "Portfolio_Decision sell-side rows (TRIM/EXIT) carry Δ Shares", "NA", None,
                         "Portfolio_Decision header not found"))
    else:
        hdr = pd_rows[pd_hdr_i]
        body = []
        for rr in pd_rows[pd_hdr_i + 1:]:
            if not rr or not _s(rr[0]):
                break
            body.append(rr)
        vals = _implied_fx(hdr, body, "Ccy", "Price", "Price SAR")
        if vals:
            outside = [(s, v) for s, v in vals if not (lo <= v <= hi)]
            out.append(Check("A2b", f"USD FX peg — Portfolio_Decision Price SAR/Price within [{lo}, {hi}]",
                             "PASS" if not outside else "FAIL", len(outside),
                             f"usd_rows={len(vals)} outside={len(outside)} "
                             f"e.g. {' '.join(f'{s}={x:.4f}' for s, x in outside[:3]) or '-'}"))
        else:
            out.append(Check("A2b", f"USD FX peg — Portfolio_Decision Price SAR/Price within [{lo}, {hi}]", "NA", None,
                             "no USD rows with both prices"))
        ix = {h: i for i, h in enumerate(hdr)}
        dcol = next((k for k in ("\u0394 Shares", "Delta Shares", "Δ Shares") if k in ix), None)
        if dcol is None or "Action" not in ix:
            out.append(Check("A4", "Portfolio_Decision sell-side rows (TRIM/EXIT) carry Δ Shares", "NA", None,
                             "Δ Shares / Action column not found"))
        else:
            sells = [r for r in body if len(r) > ix["Action"] and _s(r[ix["Action"]]).upper() in ("TRIM", "EXIT")]
            blank = [_s(r[ix["Symbol"]]) for r in sells
                     if len(r) <= ix[dcol] or _to_float(r[ix[dcol]]) is None]
            out.append(Check("A4", "Portfolio_Decision sell-side rows (TRIM/EXIT) carry Δ Shares",
                             "FAIL" if blank else "PASS", len(blank),
                             f"sell_rows={len(sells)} blank_delta_shares={' '.join(blank) or 'none'}"))
    # A3 — equity identity collisions on the equity pages
    for p in ("Global_Markets", "Market_Leaders"):
        rows = src.tab(p)
        if not rows:
            out.append(Check(f"A3-{p}", f"{p} equity identity collisions (futures / crypto / contract-month names)",
                             "NA", None, "tab unreadable"))
            continue
        hdr = [_s(c) for c in rows[0]]
        ix = {h: i for i, h in enumerate(hdr)}
        if "Symbol" not in ix or "Name" not in ix:
            out.append(Check(f"A3-{p}", f"{p} equity identity collisions (futures / crypto / contract-month names)",
                             "NA", None, "Symbol/Name columns not found"))
            continue
        hits: List[str] = []
        n_eq = 0
        for r in rows[1:]:
            if len(r) <= max(ix["Symbol"], ix["Name"]):
                continue
            sym, name = _s(r[ix["Symbol"]]), _s(r[ix["Name"]])
            if not sym or _NON_EQUITY_SYMBOL_RE.search(sym) or _symbol_suffix(sym) not in suffixes:
                continue
            n_eq += 1
            if name and _IDENTITY_COLLISION_RE.search(name):
                hits.append(f"{sym}\u2192{name[:28]}")
        out.append(Check(f"A3-{p}", f"{p} equity identity collisions (futures / crypto / contract-month names)",
                         "FAIL" if hits else "PASS", len(hits),
                         f"equity_rows={n_eq} hits={' | '.join(hits[:6]) or 'none'}"))
    return out


def run_all(src: Source, board_min: int = 5) -> List[Check]:
    checks: List[Check] = []
    for fn in (lambda: check_board(src, board_min), lambda: check_evidence_clock(src), lambda: check_learning(src),
               lambda: check_feed_truth(src), lambda: check_pages(src), lambda: check_s1(src),
               lambda: check_decision_integrity(src)):   # v1.0.6
        try:
            checks.extend(fn())
        except Exception as exc:  # a broken check is reported, never hides the rest
            checks.append(Check("ERR", f"{fn.__name__ if hasattr(fn, '__name__') else 'check'} crashed", "NA", None, str(exc)[:160]))
    checks.append(Check("D10-6", "daily brief automated (digest.yml enabled)", "NA", None,
                        "workflow state is not in the workbook — verify in Actions"))
    return checks


def render(checks: List[Check], title: str) -> str:
    w = max(len(c.name) for c in checks) + 2
    iw = max(len(c.cid) for c in checks) + 2
    lines = [f"TFB ACCEPTANCE v{VERSION} — {title} — {_now_riyadh().strftime('%Y-%m-%d %H:%M')} Riyadh",
             f"{'id':{iw}s}{'criterion':{w}s}{'verdict':8s}{'measured':>10s}  evidence"]
    for c in checks:
        m = "" if c.measured is None else (f"{c.measured:g}" if isinstance(c.measured, (int, float)) else str(c.measured)[:10])
        lines.append(f"{c.cid:{iw}s}{c.name:{w}s}{c.verdict:8s}{m:>10s}  {c.evidence[:110]}")
    tally = {v: sum(1 for c in checks if c.verdict == v) for v in ("PASS", "WARN", "FAIL", "NA")}
    lines.append("TALLY " + " ".join(f"{k}={v}" for k, v in tally.items()))
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# selftest (offline, synthetic)                                               #
# --------------------------------------------------------------------------- #
class _MemSource(Source):
    def __init__(self, tabs: Dict[str, List[List[str]]]):
        self.tabs = tabs

    def tab(self, name: str) -> List[List[str]]:
        return self.tabs.get(name, [])


def _selftest() -> int:
    now = _now_riyadh().strftime("%Y-%m-%dT%H:%M:%S+03:00")
    hdr = ["Symbol", "Name", "Open", "Day High", "Day Low", "Expected ROI 12M", "Forecast Source", "Target Price",
           "Investability Status", "Final Action", "Warnings", "Last Updated (UTC)"]
    good = [hdr, ["AAPL", "Apple", "10", "11", "9", "0.2", "provider_target", "12", "INVESTABLE", "INVEST", "", now]]
    bad = [hdr, ["Copper Futures", "Commodity", "6.49", "6.4955", "6.728", "\u25b2 31.90%", "provider_target", "", "INVESTABLE", "INVEST", "fetch_failed:HTTP 422", now],
           ["AAPL", "Apple", "10", "11", "9", "0.2", "provider_target", "", "WATCHLIST", "WATCH", "", now]]
    status_ok = [["Page", "Last Updated", "Status", "Message", "", "", "", "", "", "", "", "Global Key", "Value"],
                 ["Global_Markets", now, "PARTIAL", "data=PARTIAL guard=pw:0/2,rb:0/2", "", "", "", "", "", "", "", "TFB Decision Feed", "NOT_ACTIONABLE(partial:GM) | run=1"],
                 ["Market_Leaders", now, "SUCCESS", "data=COMPLETE guard=pw:0/2,rb:1/2", "", "", "", "", "", "", "", "", ""],
                 ["", "", "", "", "", "", "", "", "", "", "", "TFB Grid Capacity", "OK | allocated=5,000,000 (50.00%) | free=5,000,000"],
                 ["", "", "", "", "", "", "", "", "", "", "", "TFB Calibration", "INVESTABLE:0.814 | n=INVESTABLE=421"]]
    status_bad = [["Page", "Last Updated", "Status", "Message", "", "", "", "", "", "", "", "Global Key", "Value"],
                  ["Global_Markets", now, "SUCCESS", "data=PARTIAL guard=pw:0/2,rb:2/2", "", "", "", "", "", "", "", "Backend URL", "x"]]
    top10 = [["Status:", "Last run x | status: ok"], ["Deployable (SAR)", "Exp", "Selected", "Rel", "RR", "Scanned", "Passed", "Unalloc"],
             ["3218", "0", "2 / 10", "", "", "9786", "6", "0"], [], ["ALERTS (1)"], ["Type", "Count", "Action"], ["capital_call", "2", "x"], []]
    perf = [["x"], ["x"], ["x"], ["Record ID", "Key", "Symbol", "Horizon", "Date Recorded (Riyadh)", "Target Date (Riyadh)", "Status", "Outcome"],
            ["1", "K1", "AAPL", "1W", "2026-08-01", "2026-08-08 00:00:00", "matured", "WIN"],
            ["2", "K2", "AAPL", "1W", "2026-08-01", "2099-01-01 00:00:00", "active", ""]]
    runlog = [["Timestamp", "Level", "Action", "Page"], [now, "INFO", "run_dashboard_sync", "Market_Leaders"]]
    s1 = [["S-1 GATE v1.6.0", "as of x", "verdict: NOT_DECIDABLE"], ["1", "4+ weeks shadow evidence", "PENDING", "3/28 scored days"]]
    A = _MemSource({"Market_Leaders": good, "Global_Markets": good, "Commodities_FX": good, "Mutual_Funds": good,
                    "_Status": status_ok, "Top_10_Investments": top10, "Performance_Log": perf, "_Run_Log": runlog, "S1_Gate": s1})
    ca = {c.cid: c for c in run_all(A)}
    assert ca["D10-1"].verdict == "PASS" and ca["D10-1"].measured == 6
    assert ca["D10-2a"].verdict == "PASS" and ca["D10-2b"].verdict == "PASS"
    assert ca["D10-3a"].verdict == "PASS" and ca["D10-3b"].verdict == "PASS" and ca["D10-3b"].measured == 0
    assert ca["D10-4a"].verdict == "PASS" and ca["D10-4b"].verdict == "PASS"
    assert all(ca[f"G3-{p}"].verdict == "PASS" for p in PAGES)
    assert ca["D10-5-Global_Markets"].verdict == "PASS" and ca["D10-5-Market_Leaders"].verdict == "FAIL" and ca["D10-5-Market_Leaders"].measured == 1
    assert ca["D10-5-Commodities_FX"].verdict == "NA" and ca["G4-Global_Markets"].verdict == "PASS"
    # v1.0.5: a v6.56.0 tolerated stamp -> PASS with measured = survived rows; an untolerated
    # residual above 1% -> FAIL; prewrite-flagged rows never count against the write.
    status_tol = [status_ok[0],
                  ["Global_Markets", now, "SUCCESS", "data=COMPLETE rb_tol=17(+15) guard=pw:63/6609,rb:78/6609", "", "", "", "", "", "", "", "TFB Decision Feed", "EXECUTABLE | run=1"],
                  ["Market_Leaders", now, "SUCCESS", "data=COMPLETE guard=pw:30/255,rb:30/255", "", "", "", "", "", "", "", "", ""]]
    T = _MemSource({"Market_Leaders": good, "Global_Markets": good, "Commodities_FX": good, "Mutual_Funds": good,
                    "_Status": status_tol, "Top_10_Investments": top10, "Performance_Log": perf, "_Run_Log": runlog, "S1_Gate": s1})
    ct = {c.cid: c for c in run_all(T)}
    assert ct["D10-5-Global_Markets"].verdict == "PASS" and ct["D10-5-Global_Markets"].measured == 15
    assert ct["D10-5-Market_Leaders"].verdict == "PASS" and ct["D10-5-Market_Leaders"].measured == 0
    assert ca["D10-7"].verdict == "WARN" and ca["D10-7"].measured == 3
    B = _MemSource({"Market_Leaders": bad, "Global_Markets": bad, "Commodities_FX": bad, "Mutual_Funds": bad,
                    "_Status": status_bad, "Top_10_Investments": [], "Performance_Log": [], "_Run_Log": [], "S1_Gate": []})
    cb = {c.cid: c for c in run_all(B)}
    assert cb["D10-1"].verdict == "NA" and cb["D10-2a"].verdict == "FAIL" and cb["D10-2b"].verdict == "FAIL"
    assert cb["D10-3a"].verdict == "FAIL" and cb["D10-4a"].verdict == "FAIL" and cb["D10-4b"].verdict == "FAIL" and cb["D10-4b"].measured == 1
    assert cb["G3-Global_Markets"].verdict == "FAIL" and cb["G3-Global_Markets"].measured == 1
    assert cb["D10-5-Global_Markets"].verdict == "FAIL" and cb["D10-5-Global_Markets"].measured == 2 and cb["G4-Global_Markets"].verdict == "WARN" and cb["G2-Global_Markets"].verdict == "FAIL"
    assert cb["G1-Global_Markets"].verdict == "PASS"  # two distinct symbols, none blank
    # v1.0.2 fixtures (review P1/P2): stale page -> G1 FAIL even with unique symbols
    stale = [hdr, ["AAPL", "Apple", "10", "11", "9", "0.2", "provider_target", "12", "INVESTABLE", "INVEST", "", "2026-01-01T00:00:00+00:00"]]
    C = _MemSource({"Market_Leaders": stale, "Global_Markets": good, "Commodities_FX": good, "Mutual_Funds": good,
                    "_Status": status_ok, "Top_10_Investments": top10, "Performance_Log": perf, "_Run_Log": runlog, "S1_Gate": s1})
    cc = {c.cid: c for c in run_all(C)}
    assert cc["G1-Market_Leaders"].verdict == "FAIL" and cc["G1-Global_Markets"].verdict == "PASS"
    # mixed cohort: provider_target rows blank, unrelated rows carry targets -> G2 FAIL
    mixed = [hdr, ["AAPL", "Apple", "10", "11", "9", "0.2", "provider_target", "", "INVESTABLE", "INVEST", "", now],
             ["MSFT", "Microsoft", "10", "11", "9", "0.1", "phase_ii_synthetic", "50", "WATCHLIST", "WATCH", "", now]]
    D = _MemSource({"Market_Leaders": mixed, "Global_Markets": good, "Commodities_FX": good, "Mutual_Funds": good, "_Status": status_ok,
                    "Top_10_Investments": top10, "Performance_Log": perf, "_Run_Log": runlog, "S1_Gate": s1})
    cd = {c.cid: c for c in run_all(D)}
    assert cd["G2-Market_Leaders"].verdict == "FAIL" and cd["G2-Market_Leaders"].measured == 0
    # feed state criterion: NOT_ACTIONABLE is WARN (truthful), EXECUTABLE is PASS
    assert ca["D10-4c"].verdict == "WARN"
    E = _MemSource({"_Status": [["Page", "Last Updated", "Status", "Message", "", "", "", "", "", "", "", "Global Key", "Value"],
                                ["", "", "", "", "", "", "", "", "", "", "", "TFB Decision Feed", "EXECUTABLE | run=1"]]})
    assert {c.cid: c for c in check_feed_truth(E)}["D10-4c"].verdict == "PASS"
    # duplicate downloads: the NEWEST normalized export wins, deterministically, and inputs are hashed
    import tempfile, time
    with tempfile.TemporaryDirectory() as td:
        older = os.path.join(td, "_X_-_Market_Leaders(6).tsv")
        newer = os.path.join(td, "_X_-_Market_Leaders(7).tsv")
        with open(older, "w", encoding="utf-8") as fh:
            fh.write("Symbol\tName\nOLD\tx\n")
        with open(newer, "w", encoding="utf-8") as fh:
            fh.write("Symbol\tName\nNEW\tx\n")
        os.utime(older, (time.time() - 100, time.time() - 100))
        src = ExportSource(td)
        assert src.tab("Market_Leaders")[1][0] == "NEW", src.tab("Market_Leaders")
        rec = next(i for i in src.inputs if i["tab"] == "Market_Leaders")
        assert rec["file"].endswith("(7).tsv") and len(rec["sha256"]) == 64 and len(rec["candidates"]) == 2
        # missing xlsx / unreadable tab is RECORDED, never silent
        src2 = ExportSource(td, xlsx=os.path.join(td, "missing.xlsx"))
        assert any("xlsx not found" in e for e in src2.errors)
        src2.tab("Performance_Log")
        assert any("Performance_Log" in e for e in src2.errors), src2.errors
    # v1.0.6 fixtures: decision integrity A1..A4 (good and bad Top_10 board/audit, PD, GM names)
    t10_hdr_board = ["Rank", "Symbol", "Name", "Market", "Sector", "Ccy", "FX\u2192SAR", "Price", "Price SAR",
                     "Stability"]
    t10_hdr_audit = ["Symbol", "Name", "Market", "Sector", "Ccy", "Price", "Price SAR", "Verdict"]
    t10_good = [["Status:", "Last run x"], [], ["banner"], t10_hdr_board,
                ["1", "DDI.US", "DoubleDown", "NASDAQ", "Tech", "USD", "3.7528", "10.00", "37.53", "ACTIVE (2/3)"],
                ["2", "2222.SR", "Aramco", "Tadawul", "Energy", "SAR", "1.0", "25.96", "25.96", "ACTIVE (3/3)"], [],
                ["CANDIDATES \u2014 FULL AUDIT"], t10_hdr_audit] + [
                [f"S{i}.US", f"Co {i}", "NYSE", "Tech", "USD", "10.00", "37.53", "INVEST"] for i in range(6)]
    t10_bad = [["Status:", "Last run x"], [], ["banner"], t10_hdr_board,
               ["1", "HDFCBANK.NS", "HDFC Bank", "NSE India", "Financials", "INR", "0.0395", "710.30", "28.03", "GRACE"],
               ["\u2014", "BYMA.BA", "Bolsas", "BCBA", "Financials", "ARS", "", "264.50", "", "GRACE"],
               ["3", "DDI.US", "DoubleDown", "NASDAQ", "Tech", "USD", "3.8088", "10.00", "38.09", "ACTIVE"], [],
               ["CANDIDATES \u2014 FULL AUDIT"], t10_hdr_audit] + [
               [f"S{i}.US", f"Co {i}", "NYSE", "Tech", "USD", "10.00", "38.09", "INVEST"] for i in range(6)]
    pd_hdr = ["Action", "Symbol", "Name", "Sector", "Ccy", "Qty", "Avg Cost", "Price", "Price SAR", "\u0394 SAR",
              "\u0394 Shares"]
    pd_good = [["x"], pd_hdr, ["TRIM", "HCI.US", "HCI", "Fin", "USD", "30", "191.41", "189.83", "712.36", "-3,338 SAR", "-5"],
               ["HOLD", "YUM", "Yum", "CC", "USD", "24", "144.71", "152.54", "572.45", "", ""]]
    pd_bad = [["x"], pd_hdr, ["TRIM", "HCI.US", "HCI", "Fin", "USD", "30", "191.41", "189.83", "723.06", "-3,338 SAR", ""],
              ["EXIT", "OTIS", "Otis", "Ind", "USD", "55", "72.79", "71.48", "272.27", "-14,759 SAR", ""]]
    gm_hdr = ["Symbol", "Name", "Open", "Day High", "Day Low", "Expected ROI 12M", "Forecast Source", "Target Price",
              "Investability Status", "Final Action", "Warnings", "Last Updated (UTC)"]
    gm_bad = [gm_hdr, ["KE.US", "KC HRW Wheat Futures,Dec-2026", "10", "11", "9", "0.1", "provider_target", "12",
                       "WATCHLIST", "WATCH", "", now],
              ["LINK.US", "Chainlink USD", "10", "11", "9", "0.1", "provider_target", "12", "WATCHLIST", "WATCH", "", now],
              ["NG.US", "Natural Gas Oct 26", "10", "11", "9", "0.1", "provider_target", "12", "WATCHLIST", "WATCH", "", now],
              ["CC.US", "The Chemours Company", "10", "11", "9", "0.1", "provider_target", "12", "WATCHLIST", "WATCH", "", now],
              ["FTFT.US", "Future FinTech Group Inc.", "10", "11", "9", "0.1", "provider_target", "12", "WATCHLIST", "WATCH", "", now],
              ["NG=F", "Natural Gas Oct 26", "3", "3", "3", "", "", "", "", "", "", now],
              ["BTC-USD", "Bitcoin USD", "3", "3", "3", "", "", "", "", "", "", now],
              ["1120.SR", "Al Rajhi Bank", "10", "11", "9", "0.1", "provider_target", "12", "INVESTABLE", "INVEST", "", now]]
    G = _MemSource({"Top_10_Investments": t10_good, "Portfolio_Decision": pd_good, "Global_Markets": good,
                    "Market_Leaders": good})
    cg = {c.cid: c for c in check_decision_integrity(G)}
    assert cg["A1"].verdict == "PASS" and cg["A1"].measured == 0, cg["A1"].evidence
    assert cg["A2a"].verdict == "PASS" and abs(cg["A2a"].measured - 3.753) < 0.001, cg["A2a"].evidence
    assert cg["A2b"].verdict == "PASS" and cg["A2b"].measured == 0, cg["A2b"].evidence
    assert cg["A4"].verdict == "PASS" and cg["A4"].measured == 0, cg["A4"].evidence
    assert cg["A3-Global_Markets"].verdict == "PASS" and cg["A3-Market_Leaders"].verdict == "PASS"
    H = _MemSource({"Top_10_Investments": t10_bad, "Portfolio_Decision": pd_bad, "Global_Markets": gm_bad,
                    "Market_Leaders": good})
    ch = {c.cid: c for c in check_decision_integrity(H)}
    assert ch["A1"].verdict == "FAIL" and ch["A1"].measured == 2 and "HDFCBANK.NS(.NS)" in ch["A1"].evidence, ch["A1"].evidence
    assert ch["A2a"].verdict == "FAIL" and abs(ch["A2a"].measured - 3.809) < 0.001, ch["A2a"].evidence
    assert ch["A2b"].verdict == "FAIL" and ch["A2b"].measured == 2, ch["A2b"].evidence
    assert ch["A4"].verdict == "FAIL" and ch["A4"].measured == 2 and "HCI.US" in ch["A4"].evidence, ch["A4"].evidence
    assert ch["A3-Global_Markets"].verdict == "FAIL" and ch["A3-Global_Markets"].measured == 3, ch["A3-Global_Markets"].evidence
    assert "CC.US" not in ch["A3-Global_Markets"].evidence and "NG=F" not in ch["A3-Global_Markets"].evidence
    assert "FTFT.US" not in ch["A3-Global_Markets"].evidence, ch["A3-Global_Markets"].evidence   # "Future" ≠ futures
    N = _MemSource({})
    cn = {c.cid: c for c in check_decision_integrity(N)}
    assert all(cn[k].verdict == "NA" for k in ("A1", "A2a", "A2b", "A4", "A3-Global_Markets")), [c.row() for c in cn.values()]
    assert {c.cid for c in run_all(A)} >= {"A1", "A2a", "A2b", "A4", "A3-Global_Markets", "A3-Market_Leaders"}
    print(render(list(ca.values()), "selftest A (all good)"))
    print("selftest: PASS 7/7 fixtures (all-good, all-bad, stale page, mixed target cohort, duplicate downloads, unreadable xlsx, decision integrity A1..A4)")
    return 0


# --------------------------------------------------------------------------- #
def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="TFB acceptance-as-code (read-only).")
    ap.add_argument("--export-dir", default="")
    ap.add_argument("--xlsx", default="")
    ap.add_argument("--live", action="store_true")
    ap.add_argument("--sheet-id", default="")
    ap.add_argument("--json", default="")
    ap.add_argument("--board-min", type=int, default=5)
    ap.add_argument("--strict", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args(argv)
    if a.selftest:
        return _selftest()
    if a.live:
        sid = a.sheet_id or _s(os.getenv("DEFAULT_SPREADSHEET_ID")) or _s(os.getenv("SPREADSHEET_ID")) or _s(os.getenv("TARGET_SHEET_ID"))
        if not sid:
            print("FATAL: no spreadsheet id", file=sys.stderr)
            return 2
        src: Source = LiveSource(sid)
        title = f"live workbook …{sid[-6:]}"
    elif a.export_dir:
        src = ExportSource(a.export_dir, a.xlsx)
        title = f"export {os.path.basename(os.path.abspath(a.export_dir))}"
    else:
        ap.error("one of --live / --export-dir / --selftest is required")
        return 2
    checks = run_all(src, a.board_min)
    text = render(checks, title)
    print(text)
    summ = _s(os.getenv("GITHUB_STEP_SUMMARY"))
    if summ:
        try:
            with open(summ, "a", encoding="utf-8") as fh:
                fh.write("```\n" + text + "\n```\n")
        except Exception:
            pass
    errs = list(getattr(src, "errors", []) or [])
    if errs:
        print("READ ERRORS (fail-closed): " + " | ".join(errs)[:600])
    tally = {v: sum(1 for c in checks if c.verdict == v) for v in ("PASS", "WARN", "FAIL", "NA")}
    overall = "FAIL" if (tally["FAIL"] or errs) else ("WARN" if tally["WARN"] else "PASS")
    if a.json:
        os.makedirs(os.path.dirname(os.path.abspath(a.json)), exist_ok=True)
        with open(a.json, "w", encoding="utf-8") as fh:
            json.dump({"version": VERSION, "title": title, "generated_riyadh": _now_riyadh().isoformat(),
                       "provenance": _provenance(a, src),
                       "overall_verdict": overall, "fail_count": tally["FAIL"], "tally": tally,
                       "read_errors": errs, "checks": [c.row() for c in checks]}, fh, indent=2, ensure_ascii=False)
    if a.strict and (tally["FAIL"] or errs):
        return 1
    return 0


def _provenance(a, src: Source) -> Dict[str, Any]:
    """v1.0.2: who measured what, from which commit, on which inputs."""
    here = os.path.abspath(__file__)
    sha = _s(os.getenv("GITHUB_SHA"))
    if not sha:
        try:
            sha = subprocess.run(["git", "rev-parse", "HEAD"], capture_output=True, text=True, timeout=5,
                                 cwd=os.path.dirname(here)).stdout.strip()
        except Exception:
            sha = ""
    return {"generator_version": VERSION, "generator_sha256": _sha256_file(here),
            "generator_commit_sha": sha or None, "workflow_run_id": _s(os.getenv("GITHUB_RUN_ID")) or None,
            "source_mode": "live" if a.live else "export", "snapshot_riyadh": _now_riyadh().isoformat(),
            "inputs": list(getattr(src, "inputs", []) or [])}


if __name__ == "__main__":
    sys.exit(main())
