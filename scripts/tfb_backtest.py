#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tfb_backtest.py — v1.0.0 (2026-09-01)
================================================================================
WHY: the Strategy's rule is "register a hypothesis, backtest it, then change a
weight or gate". H-28 (stated reliability predicts outcomes) was rejected by an
ad-hoc computation; this makes that computation a repeatable instrument so any
column of Performance_Log can be tested the same way (H-29: which signal DOES
separate winners), by either AI, on the same export, with the same numbers.

WHAT IT DOES
  - loads Performance_Log (browser TSV export, xlsx, or --live), keeps matured
    cohorts with a WIN/LOSS outcome, DEDUPLICATED by Key (the v6.35.0 rule);
  - for each --signal column: numeric -> bands (custom edges or quintiles),
    categorical -> groups; per group n, win %, mean/median realized ROI %;
  - predictive power: Brier of the raw value as a probability (numeric 0-100
    only) vs naive 0.5 vs constant base rate; 5-fold cross-validated Brier of a
    group-calibrated probability (shrinkage k=20) vs the base rate;
    Spearman(signal, realized ROI) for numeric signals; win-rate spread
    (max-min group win %, groups with n >= min_n);
  - verdict per signal: SEPARATES (CV Brier beats base rate by >= 0.002 AND
    spread >= 5 pp with n >= min_n in both extremes) / WEAK / NONE.

USAGE
  python scripts/tfb_backtest.py --export-dir DIR [--xlsx F] --signal "Entry Forecast Reliability" --signal Confidence ...
  python scripts/tfb_backtest.py --live --signal ... (env: DEFAULT_SPREADSHEET_ID + Google creds)
  python scripts/tfb_backtest.py --export-dir DIR --all-signals      # every Entry*/Horizon/Origin column
  python scripts/tfb_backtest.py --selftest
Options: --edges "50,70,85" (numeric bands) --min-n 100 --json out.json --horizon 1W,2W,1M
         --filter "Origin Tab=Top_10_Investments" (repeatable, exact match)   [v1.1.0]
         --since 2026-08-01 / --until 2026-08-31  (Date Recorded window)       [v1.1.0]
READ-ONLY. No writes, ever.
"""
from __future__ import annotations

import argparse
import base64
import csv
import json
import math
import os
import random
import statistics
import sys
from typing import Any, Dict, List, Optional, Tuple

VERSION = "1.1.0"
DEFAULT_SIGNALS = ["Entry Forecast Reliability", "Entry Score", "Confidence", "Entry Investability",
                   "Entry Recommendation", "Entry Risk Bucket", "Horizon", "Origin Tab"]


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _f(v: Any) -> Optional[float]:
    t = _s(v).replace("%", "").replace(",", "").replace("\u25b2", "").replace("\u25bc", "").strip()
    if not t:
        return None
    try:
        x = float(t)
        return x if math.isfinite(x) else None
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# loading                                                                     #
# --------------------------------------------------------------------------- #
def _rows_from_tsv(path: str) -> List[List[str]]:
    with open(path, encoding="utf-8", newline="") as fh:
        return [list(r) for r in csv.reader(fh, delimiter="\t", quoting=csv.QUOTE_NONE)]


def _rows_from_xlsx(path: str, tab: str = "Performance_Log") -> List[List[str]]:
    from openpyxl import load_workbook  # optional dependency
    wb = load_workbook(path, read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        return []
    return [[_s(c) for c in r] for r in wb[tab].iter_rows(values_only=True)]


def _rows_live(sheet_id: str, tab: str = "Performance_Log") -> List[List[str]]:
    import gspread
    from google.oauth2 import service_account
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    path = _s(os.getenv("GOOGLE_APPLICATION_CREDENTIALS"))
    if path and os.path.exists(path):
        creds = service_account.Credentials.from_service_account_file(path, scopes=scopes)
    else:
        raw = _s(os.getenv("GOOGLE_SHEETS_CREDENTIALS")) or _s(os.getenv("GOOGLE_SHEETS_CREDENTIALS_B64"))
        if not raw.startswith("{"):
            raw = base64.b64decode(raw).decode("utf-8", errors="replace").strip()
        creds = service_account.Credentials.from_service_account_info(json.loads(raw), scopes=scopes)
    ws = gspread.authorize(creds).open_by_key(sheet_id).worksheet(tab)
    rc = int(getattr(ws, "row_count", 0) or 0)
    return [[_s(c) for c in r] for r in (ws.get(f"A1:AF{max(2, rc)}") or [])]


def load_records(rows: List[List[str]]) -> Tuple[List[str], List[Dict[str, str]]]:
    """Find the header row (first cell 'Record ID'), return (headers, records)."""
    hi = next((i for i, r in enumerate(rows[:12]) if r and _s(r[0]) == "Record ID"), None)
    if hi is None:
        return [], []
    hdr = [_s(h) for h in rows[hi]]
    out = []
    for r in rows[hi + 1:]:
        if not r or not _s(r[0]):
            continue
        out.append({h: (r[i] if i < len(r) else "") for i, h in enumerate(hdr) if h})
    return hdr, out


def decided_cohorts(records: List[Dict[str, str]], horizons: Optional[List[str]] = None,
                    filters: Optional[Dict[str, str]] = None, since: str = "", until: str = "") -> List[Dict[str, str]]:
    """Matured WIN/LOSS, one record per Key (first occurrence); optional horizon,
    exact-match column filters and a Date Recorded window (v1.1.0)."""
    seen, out = set(), []
    for r in records:
        if _s(r.get("Status")).lower() != "matured" or _s(r.get("Outcome")) not in ("WIN", "LOSS"):
            continue
        # dedup FIRST: the canonical cohort for a Key is its first occurrence,
        # whatever window or filter is applied afterwards (v1.1.0 fix).
        k = _s(r.get("Key"))
        if k in seen:
            continue
        seen.add(k)
        if horizons and _s(r.get("Horizon")) not in horizons:
            continue
        if filters and any(_s(r.get(k2)) != v for k2, v in filters.items()):
            continue
        d = _s(r.get("Date Recorded (Riyadh)"))[:10]
        if since and d and d < since:
            continue
        if until and d and d > until:
            continue
        if _f(r.get("Realized ROI %")) is None:
            continue
        out.append(r)
    return out


# --------------------------------------------------------------------------- #
# statistics (pure)                                                           #
# --------------------------------------------------------------------------- #
def brier(pairs: List[Tuple[float, float]]) -> float:
    return sum((p - y) ** 2 for p, y in pairs) / len(pairs) if pairs else float("nan")


def spearman(a: List[float], b: List[float]) -> float:
    n = len(a)
    if n < 3:
        return float("nan")
    def ranks(x):
        order = sorted(range(n), key=lambda i: x[i])
        r = [0.0] * n
        i = 0
        while i < n:
            j = i
            while j + 1 < n and x[order[j + 1]] == x[order[i]]:
                j += 1
            avg = (i + j) / 2.0 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r
    ra, rb = ranks(a), ranks(b)
    d2 = sum((ra[i] - rb[i]) ** 2 for i in range(n))
    return 1 - 6 * d2 / (n * (n * n - 1))


def cv_brier(groups: List[Any], ys: List[float], k: int = 5, shrink: float = 20.0, seed: int = 7) -> float:
    idx = list(range(len(ys)))
    random.Random(seed).shuffle(idx)
    folds = [set(idx[i::k]) for i in range(k)]
    out = []
    for f in folds:
        train = [i for i in idx if i not in f]
        base = sum(ys[i] for i in train) / max(1, len(train))
        agg: Dict[Any, List[float]] = {}
        for i in train:
            a = agg.setdefault(groups[i], [0.0, 0.0])
            a[0] += ys[i]
            a[1] += 1
        for i in f:
            s, n = agg.get(groups[i], [0.0, 0.0])
            p = (s + base * shrink) / (n + shrink) if n else base
            out.append((p, ys[i]))
    return brier(out)


def evaluate_signal(cohorts: List[Dict[str, str]], signal: str, edges: Optional[List[float]] = None,
                    min_n: int = 100) -> Dict[str, Any]:
    vals = [r.get(signal) for r in cohorts]
    nums = [_f(v) for v in vals]
    numeric = sum(1 for x in nums if x is not None) >= 0.9 * max(1, len(vals))
    ys = [1.0 if _s(r.get("Outcome")) == "WIN" else 0.0 for r in cohorts]
    rois = [_f(r.get("Realized ROI %")) or 0.0 for r in cohorts]
    if numeric:
        keep = [i for i, x in enumerate(nums) if x is not None]
        xs = [nums[i] for i in keep]
        if edges is None:  # quintiles
            qs = sorted(xs)
            cuts = [qs[int(len(qs) * q)] for q in (0.2, 0.4, 0.6, 0.8)]
            label = lambda x: "Q%d" % (1 + sum(1 for c in cuts if x >= c))
        else:
            e = sorted(edges)
            def label(x):
                for j, c in enumerate(e):
                    if x < c:
                        return f"<{c:g}" if j == 0 else f"{e[j-1]:g}-{c:g}"
                return f">={e[-1]:g}"
        groups = [label(nums[i]) for i in keep]
    else:
        keep = [i for i, v in enumerate(vals) if _s(v)]
        groups = [_s(vals[i]) for i in keep]
    y = [ys[i] for i in keep]
    roi = [rois[i] for i in keep]
    per: Dict[str, Dict[str, Any]] = {}
    for g, yy, rr in zip(groups, y, roi):
        d = per.setdefault(g, {"n": 0, "wins": 0.0, "rois": []})
        d["n"] += 1
        d["wins"] += yy
        d["rois"].append(rr)
    table = []
    for g in sorted(per, key=lambda k: (-per[k]["n"] if not numeric else 0, k)):
        d = per[g]
        table.append({"group": g, "n": d["n"], "win_pct": round(100 * d["wins"] / d["n"], 1),
                      "mean_roi_pct": round(statistics.mean(d["rois"]), 2),
                      "median_roi_pct": round(statistics.median(d["rois"]), 2)})
    base = sum(y) / max(1, len(y))
    base_brier = brier([(base, yy) for yy in y])
    res: Dict[str, Any] = {"signal": signal, "type": "numeric" if numeric else "categorical", "n": len(y),
                           "base_win_pct": round(100 * base, 1), "base_brier": round(base_brier, 4), "groups": table}
    if numeric:
        xs = [nums[i] for i in keep]
        if xs and 0 <= min(xs) and max(xs) <= 100:
            res["raw_brier_as_probability"] = round(brier([(x / 100.0, yy) for x, yy in zip(xs, y)]), 4)
        res["spearman_vs_roi"] = round(spearman(xs, roi), 3)
    res["cv_brier_group_calibrated"] = round(cv_brier(groups, y), 4)
    big = [t for t in table if t["n"] >= min_n]
    spread, z = 0.0, 0.0
    if len(big) >= 2:
        hi = max(big, key=lambda t: t["win_pct"]); lo = min(big, key=lambda t: t["win_pct"])
        spread = hi["win_pct"] - lo["win_pct"]
        p = base
        se = math.sqrt(max(1e-9, p * (1 - p) * (1.0 / hi["n"] + 1.0 / lo["n"])))
        z = (spread / 100.0) / se
    res["win_spread_pp"] = round(spread, 1)
    res["spread_z"] = round(z, 2)
    gain = base_brier - res["cv_brier_group_calibrated"]
    res["cv_gain_vs_base"] = round(gain, 4)
    # SEPARATES: out-of-sample Brier gain AND a spread that is not chance (z >= 3);
    # WEAK: one of the two; NONE: neither. Chance spreads across quintiles never reach z >= 3.
    strong_gain, strong_spread = gain >= 0.002, (z >= 3.0 and spread >= 5.0)
    res["verdict"] = "SEPARATES" if (strong_gain and strong_spread) else ("WEAK" if (strong_gain or strong_spread) else "NONE")
    return res


def render(results: List[Dict[str, Any]], title: str) -> str:
    lines = [f"TFB BACKTEST v{VERSION} — {title}"]
    for r in results:
        lines.append(f"\n[{r['verdict']:9s}] {r['signal']} ({r['type']}, n={r['n']}, base win {r['base_win_pct']}%, base Brier {r['base_brier']})")
        if "raw_brier_as_probability" in r:
            lines.append(f"    raw value as probability: Brier {r['raw_brier_as_probability']} | Spearman vs ROI {r['spearman_vs_roi']}")
        lines.append(f"    CV Brier group-calibrated {r['cv_brier_group_calibrated']} (gain vs base {r['cv_gain_vs_base']:+.4f}) | win spread {r['win_spread_pp']} pp (z={r['spread_z']})")
        lines.append(f"    {'group':14s}{'n':>6s}{'win%':>7s}{'meanROI%':>10s}{'medROI%':>9s}")
        for t in r["groups"][:12]:
            lines.append(f"    {t['group'][:14]:14s}{t['n']:6d}{t['win_pct']:7.1f}{t['mean_roi_pct']:10.2f}{t['median_roi_pct']:9.2f}")
    return "\n".join(lines)


def _selftest() -> int:
    random.seed(1)
    recs = []
    for i in range(3000):
        sig = random.uniform(0, 100)            # informative signal: P(win) rises with it
        noise = random.uniform(0, 100)          # uninformative
        win = random.random() < 0.35 + 0.5 * (sig / 100.0)
        recs.append({"Key": f"K{i}", "Status": "matured", "Outcome": "WIN" if win else "LOSS",
                     "Realized ROI %": str(round((random.gauss(2, 6) if win else random.gauss(-3, 5)), 2)),
                     "Good": str(round(sig, 1)), "Noise": str(round(noise, 1)),
                     "Cat": "A" if sig > 60 else "B", "Horizon": "1W"})
    recs.append(dict(recs[0], Key="K0"))        # duplicate key must be dropped
    coh = decided_cohorts(recs)
    assert len(coh) == 3000, len(coh)
    assert len(decided_cohorts(recs, filters={"Cat": "A"})) == sum(1 for r in recs[:3000] if r["Cat"] == "A")
    for r in recs[:1500]:
        r["Date Recorded (Riyadh)"] = "2026-07-15 09:00:00"
    for r in recs[1500:]:
        r["Date Recorded (Riyadh)"] = "2026-08-20 09:00:00"
    assert len(decided_cohorts(recs, since="2026-08-01")) == 1500 and len(decided_cohorts(recs, until="2026-07-31")) == 1500
    good = evaluate_signal(coh, "Good", edges=[50, 70, 85]); noise = evaluate_signal(coh, "Noise")
    cat = evaluate_signal(coh, "Cat")
    assert good["verdict"] == "SEPARATES" and good["win_spread_pp"] >= 20, good
    assert noise["verdict"] == "NONE", noise["verdict"]
    assert cat["verdict"] == "SEPARATES" and cat["type"] == "categorical", cat
    assert good["raw_brier_as_probability"] < noise["raw_brier_as_probability"]
    print(render([good, noise, cat], "selftest"))
    print("selftest: PASS 4/4 (informative numeric SEPARATES, noise NONE, categorical SEPARATES; duplicate key dropped; filter/since/until)")
    return 0


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Cohort-outcome backtester for hypothesis-registry items (read-only).")
    ap.add_argument("--export-dir", default="")
    ap.add_argument("--xlsx", default="")
    ap.add_argument("--live", action="store_true")
    ap.add_argument("--sheet-id", default="")
    ap.add_argument("--signal", action="append", default=[])
    ap.add_argument("--all-signals", action="store_true")
    ap.add_argument("--edges", default="")
    ap.add_argument("--horizon", default="")
    ap.add_argument("--filter", action="append", default=[], help='COL=VALUE exact match, repeatable (v1.1.0)')
    ap.add_argument("--since", default="", help="Date Recorded >= YYYY-MM-DD (v1.1.0)")
    ap.add_argument("--until", default="", help="Date Recorded <= YYYY-MM-DD (v1.1.0)")
    ap.add_argument("--min-n", type=int, default=100)
    ap.add_argument("--json", default="")
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args(argv)
    if a.selftest:
        return _selftest()
    rows: List[List[str]] = []
    if a.live:
        sid = a.sheet_id or _s(os.getenv("DEFAULT_SPREADSHEET_ID")) or _s(os.getenv("SPREADSHEET_ID"))
        rows = _rows_live(sid)
        title = f"live …{sid[-6:]}"
    else:
        d = a.export_dir
        tsv = next((os.path.join(d, fn) for fn in sorted(os.listdir(d)) if fn.endswith(".tsv") and "Performance_Log" in fn), "") if d else ""
        xl = a.xlsx or (next((os.path.join(d, fn) for fn in sorted(os.listdir(d)) if fn.lower().endswith(".xlsx")), "") if d else "")
        if tsv:
            rows = _rows_from_tsv(tsv); title = os.path.basename(tsv)
        elif xl:
            rows = _rows_from_xlsx(xl); title = os.path.basename(xl)
        else:
            ap.error("no Performance_Log TSV / xlsx found; use --live, --export-dir or --xlsx")
    hdr, recs = load_records(rows)
    if not recs:
        print("FATAL: Performance_Log header row not found / no records", file=sys.stderr)
        return 2
    hz = [h.strip() for h in a.horizon.split(",") if h.strip()] or None
    flt = {}
    for item in a.filter:
        if "=" in item:
            k, v = item.split("=", 1)
            flt[k.strip()] = v.strip()
    coh = decided_cohorts(recs, hz, flt or None, a.since.strip(), a.until.strip())
    signals = a.signal or (DEFAULT_SIGNALS if a.all_signals else ["Entry Forecast Reliability"])
    edges = [float(x) for x in a.edges.split(",") if x.strip()] or None
    results = [evaluate_signal(coh, sg, edges=edges if sg == signals[0] and edges else None, min_n=a.min_n) for sg in signals if sg in hdr]
    missing = [sg for sg in signals if sg not in hdr]
    scope = (f" | horizons {','.join(hz)}" if hz else "") + (f" | filter {flt}" if flt else "") + \
            (f" | since {a.since}" if a.since else "") + (f" | until {a.until}" if a.until else "")
    print(render(results, f"{title} | decided cohorts n={len(coh)}{scope}"))
    if missing:
        print("signals not in header:", ", ".join(missing))
    if a.json:
        os.makedirs(os.path.dirname(os.path.abspath(a.json)), exist_ok=True)
        with open(a.json, "w", encoding="utf-8") as fh:
            json.dump({"version": VERSION, "title": title, "n": len(coh), "results": results, "missing": missing}, fh, indent=2, ensure_ascii=False)
    return 0


if __name__ == "__main__":
    sys.exit(main())
